import openpyxl.styles
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import shutil
from utils import setup_logging
from xlsx_to_kml import parse_coordinates, create_kml_from_coordinates
import simplekml


class WaterObjectsProcessor:
    def __init__(self, input_file, output_dir):
        self.input_file = input_file
        self.output_dir = os.path.join(output_dir, "separated_regions")
        self.history_file = os.path.join(output_dir, "processing_history.json")
        self.changes_dir = os.path.join(output_dir, "changes_reports")
        self.logger = setup_logging(output_dir)
        self.logger.info("Инициализация WaterObjectsProcessor")
        self.logger.info(f"Входной файл: {self.input_file}")
        self.logger.info(f"Директория для разделенных регионов: {self.output_dir}")
        self.logger.info(f"Файл истории: {self.history_file}")
        self.logger.info(f"Директория отчетов об изменениях: {self.changes_dir}")

    def get_column_widths(self, input_file):
        """
        Получает ширину столбцов из исходного файла.
        """
        self.logger.info(f"Получение ширины столбцов из файла: {input_file}")
        try:
            wb = load_workbook(input_file)
            ws = wb.active
            column_widths = []
            for i in range(1, ws.max_column + 1):
                column_letter = get_column_letter(i)
                width = ws.column_dimensions[column_letter].width
                column_widths.append(width)
            self.logger.debug(f"Получены ширины столбцов: {column_widths}")
            return column_widths
        except Exception as e:
            self.logger.error(f"Ошибка при получении ширины столбцов: {e}")
            return None

    def apply_formatting(self, filename, column_widths):
        """
        Применяет форматирование к файлу.

        Args:
            filename (str): Путь к Excel-файлу
            column_widths (list): Список значений ширины столбцов
        """
        self.logger.info(f"Применение форматирования к файлу: {filename}")
        try:
            wb = load_workbook(filename)
            ws = wb.active

            # Применяем ширину столбцов
            for i, width in enumerate(column_widths, 1):
                if width is not None:
                    column_letter = get_column_letter(i)
                    ws.column_dimensions[column_letter].width = width

            # Объединяем B1 и C1
            ws.merge_cells('B1:C1')
            self.logger.debug("Объединены ячейки B1:C1")

            # Объединяем ячейки в столбцах A, D, E, F, G
            columns_to_merge = ['A', 'D', 'E', 'F', 'G']
            for col in columns_to_merge:
                ws.merge_cells(f'{col}1:{col}2')
            self.logger.debug(
                f"Объединены ячейки в столбцах: {columns_to_merge}")

            # Объединяем ячейки в строке 3
            ws.merge_cells('A3:G3')
            self.logger.debug("Объединены ячейки A3:G3")

            # Устанавливаем перенос текста
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

            # Центрируем текст
            merged_cell = ws['A3']
            merged_cell.alignment = openpyxl.styles.Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

            merged_cell = ws['B1']
            merged_cell.alignment = openpyxl.styles.Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

            wb.save(filename)
            self.logger.info(
                f"Форматирование успешно применено к файлу: {filename}")
        except Exception as e:
            self.logger.error(
                f"Ошибка при применении форматирования к файлу {filename}: {e}")

    def load_processing_history(self):
        """Загружает историю обработки файлов"""
        self.logger.info("Загрузка истории обработки")
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
                self.logger.info(
                    f"Загружена история обработки: {len(history)} записей")
                return history
            except Exception as e:
                self.logger.error(f"Ошибка при загрузке истории: {e}")
                return {}
        self.logger.info("История обработки не найдена, создается новая")
        return {}

    def save_processing_history(self, history):
        """Сохраняет историю обработки файлов"""
        self.logger.info("Сохранение истории обработки")
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=4)
            self.logger.info("История обработки успешно сохранена")
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении истории: {e}")

    def get_regions_data(self, df):
        """
        Извлекает данные по регионам из DataFrame
        Returns:
            dict: {region_n,ame: region_df}
        """
        regions_data = {}
        header_rows = df.iloc[3:5]

        start_row = None
        current_region = None

        for i, row in df.iterrows():
            if i < 6:
                continue

            row_value = row.iloc[0]
            if pd.notna(row_value) and isinstance(row_value, str) and "область" in row_value:
                if start_row is not None:
                    region_df = df.iloc[start_row:i]
                    region_df = pd.concat(
                        [header_rows, region_df], ignore_index=True)
                    regions_data[current_region] = region_df

                start_row = i
                current_region = row_value.strip()

            elif pd.isna(row_value) and start_row is not None and i > 7:
                region_df = df.iloc[start_row:i]
                region_df = pd.concat(
                    [header_rows, region_df], ignore_index=True)
                regions_data[current_region] = region_df
                start_row = None

        if start_row is not None:
            region_df = df.iloc[start_row:]
            region_df = pd.concat([header_rows, region_df], ignore_index=True)
            regions_data[current_region] = region_df

        return regions_data

    def compare_regions(self, current_df, previous_df):
        """
        Сравнивает данные водных объектов в регионе между версиями

        Args:
            current_df: DataFrame текущей версии региона
            previous_df: DataFrame предыдущей версии региона

        Returns:
            tuple: (has_changes, new_objects, removed_objects, modified_objects)
        """
        try:
            # Пропускаем заголовочные строки и берем только данные объектов
            current_data = current_df.iloc[2:]
            previous_data = previous_df.iloc[2:]

            # Создаем уникальные идентификаторы для водных объектов
            # Используем комбинацию названия, местоположения и цели использования
            def create_object_id(row):
                # колонки B, C, D
                return tuple(str(x).strip() for x in row.iloc[[1, 2, 3]])

            current_objects = {create_object_id(
                row): row for _, row in current_data.iterrows()}
            previous_objects = {create_object_id(
                row): row for _, row in previous_data.iterrows()}

            current_ids = set(current_objects.keys())
            previous_ids = set(previous_objects.keys())

            # Находим новые объекты
            new_objects = []
            for obj_id in (current_ids - previous_ids):
                new_objects.append({
                    'id': obj_id,
                    'data': current_objects[obj_id].to_dict()
                })

            # Находим удаленные объекты
            removed_objects = []
            for obj_id in (previous_ids - current_ids):
                removed_objects.append({
                    'id': obj_id,
                    'data': previous_objects[obj_id].to_dict()
                })

            # Находим измененные объекты
            modified_objects = []
            for obj_id in (current_ids & previous_ids):
                current_obj = current_objects[obj_id]
                previous_obj = previous_objects[obj_id]

                # Сравниваем все значения в строках
                if not current_obj.equals(previous_obj):
                    modified_objects.append({
                        'id': obj_id,
                        'old_data': previous_obj.to_dict(),
                        'new_data': current_obj.to_dict()
                    })

            has_changes = bool(
                new_objects or removed_objects or modified_objects)
            return has_changes, new_objects, removed_objects, modified_objects

        except Exception as e:
            self.logger.error(f"Ошибка при сравнении данных региона: {e}")
            return False, [], [], []

    def save_changes_report(self, region_name, new_objects, removed_objects, modified_objects):
        """
        Сохраняет подробный отчет об изменениях водных объектов
        """
        os.makedirs(self.changes_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = os.path.join(
            self.changes_dir, f"{region_name}_changes_{timestamp}.txt")

        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(
                f"Отчет об изменениях водных объектов для региона: {region_name}\n")
            f.write(
                f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            # Новые объекты
            f.write(f"Новые водные объекты ({len(new_objects)}):\n")
            for obj in new_objects:
                f.write("-" * 50 + "\n")
                f.write(f"Название: {obj['data'].get(1, 'Н/Д')}\n")
                f.write(f"Местоположение: {obj['data'].get(2, 'Н/Д')}\n")
                f.write(f"Цель использования: {obj['data'].get(3, 'Н/Д')}\n")

            # Удаленные объекты
            f.write(f"\nУдаленные водные объекты ({len(removed_objects)}):\n")
            for obj in removed_objects:
                f.write("-" * 50 + "\n")
                f.write(f"Название: {obj['data'].get(1, 'Н/Д')}\n")
                f.write(f"Местоположение: {obj['data'].get(2, 'Н/Д')}\n")
                f.write(f"Цель использования: {obj['data'].get(3, 'Н/Д')}\n")

            # Измененные объекты
            f.write(
                f"\nИзмененные водные объекты ({len(modified_objects)}):\n")
            for obj in modified_objects:
                f.write("-" * 50 + "\n")
                f.write(f"Водный объект: {obj['id'][0]}\n")
                f.write("Изменения:\n")

                old_data = obj['old_data']
                new_data = obj['new_data']

                # Сравниваем значения по всем колонкам
                for col in old_data.keys():
                    old_val = str(old_data[col]).strip()
                    new_val = str(new_data[col]).strip()
                    if old_val != new_val:
                        f.write(f"Колонка {col}:\n")
                        f.write(f"  Было:  {old_val}\n")
                        f.write(f"  Стало: {new_val}\n")

    def save_anomalous_xlsx(self, region_name, anomalous_df, header_rows, column_widths, anomalous_xlsx_dir):
        """Сохраняет аномальные данные региона в отдельный XLSX файл."""
        if anomalous_df.empty:
            return

        os.makedirs(anomalous_xlsx_dir, exist_ok=True)
        safe_region_name = region_name.strip().replace(' ', '_').replace('/', '_') # Санитизация имени файла
        output_filename = os.path.join(anomalous_xlsx_dir, f"ANO_{safe_region_name}.xlsx")

        try:
            # Добавляем заголовки обратно
            final_df = pd.concat([header_rows, anomalous_df], ignore_index=True)

            # Сохраняем в Excel
            final_df.to_excel(output_filename, index=False, header=False)
            self.logger.info(f"Аномальные данные для региона '{region_name}' сохранены в {output_filename}")

            # Применяем форматирование (если ширина столбцов известна)
            if column_widths is not None:
                # Добавляем ширину для новой колонки "Причина аномалии"
                # Предполагаем, что она последняя
                max_col_index = final_df.shape[1] - 1
                # Используем ширину предпоследней колонки или значение по умолчанию
                last_col_width = column_widths.get(max_col_index -1, 20) if max_col_index > 0 else 20
                extended_column_widths = {**column_widths, max_col_index: last_col_width}
                self.apply_formatting(output_filename, extended_column_widths)
            else:
                # Если исходной ширины нет, применяем только объединение ячеек заголовка
                self.apply_formatting(output_filename, None)

        except Exception as e:
            self.logger.error(f"Ошибка при сохранении аномального XLSX файла {output_filename}: {e}")
            raise # Передаем исключение выше для логгирования в process_file

    def process_file(self):
        """Основной метод обработки файла"""
        self.logger.info(f"Начало обработки файла: {self.input_file}")
        history = self.load_processing_history()

        try:
            # Получаем ширину столбцов из исходного файла
            column_widths = self.get_column_widths(self.input_file)
            if column_widths is None:
                self.logger.warning(
                    "Не удалось получить ширину столбцов, продолжаем без форматирования")

            # Загружаем текущий файл
            current_df = pd.read_excel(self.input_file, header=None)
            self.logger.info(
                f"Загружен текущий файл. Размер: {current_df.shape}")

            # Получаем данные по регионам из текущего файла
            all_regions_data = self.get_regions_data(current_df)
            self.logger.info(f"Найдено регионов: {len(all_regions_data)}")

            # Создаем директорию для выходных файлов регионов и KML
            os.makedirs(self.output_dir, exist_ok=True)
            kml_output_dir = os.path.join(self.output_dir, "kml")
            os.makedirs(kml_output_dir, exist_ok=True)
            # Директории для аномальных данных (создаются в функциях сохранения)
            anomalous_xlsx_dir = os.path.join(self.output_dir, "anomalous")
            anomalous_kml_dir = os.path.join(kml_output_dir, "anomalous")

            # Загружаем предыдущую версию файла
            previous_file = history.get('latest_file')
            if previous_file and os.path.exists(previous_file):
                previous_df = pd.read_excel(previous_file, header=None)
                previous_regions_data = self.get_regions_data(previous_df)
                self.logger.info("Загружена предыдущая версия файла")
            else:
                previous_regions_data = {}
                self.logger.info("Предыдущая версия файла не найдена")

            changed_regions_count = 0
            new_regions_count = 0
            processed_regions_count = 0

            # Обрабатываем каждый регион
            for region_name, current_region_df in all_regions_data.items():
                processed_regions_count += 1
                self.logger.info(f"Обработка региона: {region_name} ({processed_regions_count}/{len(all_regions_data)})")

                # ---- Разделение на валидные и аномальные данные (Task 2.1, 2.2) ----
                if current_region_df.shape[0] <= 2: # Только заголовок
                    self.logger.info(f"Регион '{region_name}' не содержит данных, пропуск.")
                    continue

                header_rows = current_region_df.iloc[0:2]
                data_rows_df = current_region_df.iloc[2:].copy() # Работаем с копией
                coord_col_index = 4 # Индекс колонки "Место водопользования" (E), 0-based

                valid_rows = []
                anomalous_rows = []

                for index, row in data_rows_df.iterrows():
                    coord_str = row.iloc[coord_col_index]
                    anomaly_reason = None

                    # Проверка 1: Пустые координаты?
                    if pd.isna(coord_str) or not str(coord_str).strip():
                        anomaly_reason = "Отсутствуют координаты"
                    else:
                        coord_str = str(coord_str).strip()
                        # Проверка 2: Есть ли маркеры?
                        if 'м.' not in coord_str and '°' not in coord_str:
                             # Не считаем ошибкой, если строка явно не похожа на координаты
                             # Это может быть адрес.
                             # parse_coordinates сама вернет [], None
                             pass # Пропускаем к parse_coordinates

                        # Проверка 3: Парсинг и валидация
                        parsed_coords, error_reason = parse_coordinates(coord_str)

                        if error_reason is not None:
                            # parse_coordinates вернула ошибку
                            anomaly_reason = error_reason
                        elif parsed_coords is None:
                             # Не должно происходить при новой логике parse_coordinates,
                             # но на всякий случай
                             anomaly_reason = "Неизвестная ошибка парсинга"
                        # else: # Успешный парсинг (parsed_coords - список, возможно пустой)
                            # Строка валидна с точки зрения формата/диапазона

                    if anomaly_reason:
                        row_copy = row.copy()
                        # Добавляем причину аномалии (можно добавить как последнюю колонку)
                        # row_copy['Причина аномалии'] = anomaly_reason
                        anomalous_rows.append((row_copy, anomaly_reason))
                    else:
                        valid_rows.append(row)
                # ---- Конец разделения ----

                # ---- Создание DataFrame'ов ----
                valid_df = pd.DataFrame(valid_rows)
                if anomalous_rows:
                    anomalous_data, reasons = zip(*anomalous_rows)
                    anomalous_df = pd.DataFrame(list(anomalous_data))
                    anomalous_df['Причина аномалии'] = list(reasons)
                else:
                    anomalous_df = pd.DataFrame(columns=data_rows_df.columns.tolist() + ['Причина аномалии'])

                self.logger.info(f"Регион '{region_name}': Найдено валидных строк: {len(valid_df)}, аномальных: {len(anomalous_df)}")

                # ---- Сохранение аномальных данных (Task 2.3, 2.4) ----
                if not anomalous_df.empty:
                    try:
                        self.save_anomalous_xlsx(region_name, anomalous_df, header_rows, column_widths, anomalous_xlsx_dir)
                    except Exception as e:
                        self.logger.error(f"Ошибка при сохранении ANO_*.xlsx для региона '{region_name}': {e}")
                # ---- Конец сохранения аномальных ----

                # ---- Обработка валидных данных ----
                if valid_df.empty:
                    self.logger.info(f"В регионе '{region_name}' не найдено валидных строк для дальнейшей обработки.")
                    # Если и аномальных нет, регион пуст. Если были аномальные, они сохранены.
                    # Если регион был в previous_regions_data, он считается удаленным/очищенным?
                    # Текущая логика сравнения ниже обработает это как изменение (все старые удалены).
                    # Создаем DataFrame с заголовками, чтобы сравнение работало
                    current_valid_region_df = header_rows.copy()
                else:
                    # Собираем валидный DataFrame с заголовками
                    current_valid_region_df = pd.concat([header_rows, valid_df], ignore_index=True)

                # ---- Сравнение и сохранение ИЗМЕНЕННЫХ валидных данных ----
                region_changed = False
                region_is_new = False
                if region_name in previous_regions_data:
                    # Сравниваем только валидные данные
                    has_changes, new_objects, removed_objects, modified_objects = self.compare_regions(
                        current_valid_region_df, previous_regions_data[region_name]
                    )
                    if has_changes:
                        self.logger.info(f"Обнаружены изменения в ВАЛИДНЫХ данных региона: {region_name}")
                        changed_regions_count += 1
                        region_changed = True
                        # Сохраняем отчет об изменениях (только для валидных данных)
                        self.save_changes_report(
                            region_name, new_objects, removed_objects, modified_objects)
                    else:
                        self.logger.info(f"Валидные данные в регионе '{region_name}' не изменились.")
                else:
                    # Новый регион
                    self.logger.info(f"Обнаружен новый регион: {region_name}")
                    new_regions_count += 1
                    region_is_new = True
                    # Считаем новый регион "измененным", если он не пустой
                    region_changed = not current_valid_region_df.iloc[2:].empty

                # Сохраняем .xlsx и .kml для региона, ТОЛЬКО если он новый или изменился,
                # и ТОЛЬКО если в нем есть валидные данные
                if region_changed and not valid_df.empty:
                    self.logger.info(f"Сохранение валидных данных для региона: {region_name}")
                    output_xlsx_file = os.path.join(self.output_dir,
                                                f"{region_name.strip().replace(' ', '_')}.xlsx")
                    output_kml_file = os.path.join(kml_output_dir,
                                                f"{region_name.strip().replace(' ', '_')}.kml")

                    # Сохраняем XLSX
                    try:
                        current_valid_region_df.to_excel(output_xlsx_file, index=False, header=False)
                        # Применяем форматирование XLSX
                        if column_widths is not None:
                             self.apply_formatting(output_xlsx_file, column_widths)
                    except Exception as e:
                        self.logger.error(f"Ошибка при сохранении XLSX для региона '{region_name}': {e}")

                    # Создаем KML из валидного DataFrame
                    try:
                        # Создаем временный Workbook в памяти для create_kml_from_coordinates
                        # Это обходной путь, т.к. create_kml_from_coordinates ожидает openpyxl sheet
                        # TODO: Рефакторинг create_kml_from_coordinates для приема DataFrame
                        with pd.ExcelWriter('temp_kml_sheet.xlsx', engine='openpyxl') as writer:
                            current_valid_region_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
                        wb = load_workbook('temp_kml_sheet.xlsx')
                        sheet = wb.active
                        create_kml_from_coordinates(sheet, output_kml_file)
                        os.remove('temp_kml_sheet.xlsx') # Удаляем временный файл
                        self.logger.info(f"KML файл для региона '{region_name}' создан: {output_kml_file}")
                    except Exception as e:
                        self.logger.error(f"Ошибка при создании KML для региона '{region_name}': {e}")
                        if os.path.exists('temp_kml_sheet.xlsx'):
                            os.remove('temp_kml_sheet.xlsx')
                elif not valid_df.empty:
                     self.logger.info(f"Валидные данные региона '{region_name}' не изменились, сохранение пропущено.")

                # ---- Конец обработки валидных данных ----

            # --- Конец цикла по регионам ---

            self.logger.info(f"Обработка файла завершена. Всего регионов: {processed_regions_count}, "
                             f"Новых: {new_regions_count}, Измененных (валидные данные): {changed_regions_count}")

            # Обновляем историю
            history['latest_file'] = self.input_file
            history['last_processed'] = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S")
            # В историю записываем все обработанные регионы?
            # Или только те, у кого изменились валидные данные?
            # Давайте записывать все, где были изменения (включая только аномальные)
            # history['changed_regions'] = changed_regions # Старая логика
            # Собираем список регионов, где были хоть какие-то изменения или аномалии
            updated_processed_regions = list(all_regions_data.keys()) # Placeholder
            # TODO: Уточнить, что должно попадать в history['changed_regions']
            history['processed_regions'] = updated_processed_regions # Записываем все обработанные

            # Сохраняем копию текущего файла в историю
            history_dir = os.path.join("input", "history") # Изменен путь
            os.makedirs(history_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            history_file = os.path.join(
                history_dir, f"water_objects_{timestamp}.xlsx")
            shutil.copy2(self.input_file, history_file)

            self.save_processing_history(history)

            self.logger.info(
                f"Обработка завершена. Изменения найдены в {len(changed_regions_count)} регионах")
            if changed_regions_count:
                self.logger.info("Измененные регионы:")
                for region in updated_processed_regions:
                    self.logger.info(f"- {region}")

        except Exception as e:
            self.logger.error(f"Ошибка при обработке файла: {e}")
            raise


# Пример использования
if __name__ == "__main__":
    input_excel_file = "input/Объекты.xlsx"
    output_directory = "output/separated_regions"

    processor = WaterObjectsProcessor(input_excel_file, output_directory)
    processor.process_file()
