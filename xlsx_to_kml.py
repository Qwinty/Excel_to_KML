import glob
import json
import logging
import re
import math
import os
from typing import List, Tuple, Optional
import simplekml
from openpyxl import Workbook
from utils import setup_logging


from pyproj import CRS, Transformer

# Import necessary functions and the setup function from utils
from utils import generate_random_color, sort_coordinates, setup_logging

# Set up logging
logger = setup_logging()


def create_transformer(proj4_str: str) -> Transformer:
    """Создает трансформер из заданной строки Proj4 в WGS84."""
    crs = CRS.from_proj4(proj4_str)
    return Transformer.from_crs(crs, "EPSG:4326", always_xy=True)


# Определяем строки Proj4 и создаем трансформеры
try:
    with open("data/proj4.json", "r", encoding="utf-8") as f:
        proj4_strings = json.load(f)
    # Создаем трансформеры
    transformers = {name: create_transformer(proj4) for name, proj4 in proj4_strings.items()}
except FileNotFoundError:
    logger.critical("Critical Error: Could not find 'data/proj4.json'. This file is required for coordinate transformations. Ensure it exists in the 'data' directory relative to the application.")
    print("[bold red]Критическая ошибка: Не найден файл 'data/proj4.json'.[/bold red]")
    print("[bold red]Этот файл необходим для преобразования координат. Убедитесь, что он находится в папке 'data' рядом с программой.[/bold red]")
    # Exit or raise a custom exception if the program cannot function without it
    raise SystemExit("Missing essential data file: data/proj4.json")
except json.JSONDecodeError:
    logger.critical(
        "Critical Error: Could not parse 'data/proj4.json'. Check the file format.")
    print("[bold red]Критическая ошибка: Не удалось прочитать файл 'data/proj4.json'. Проверьте формат файла.[/bold red]")
    raise SystemExit("Invalid format for essential data file: data/proj4.json")
except Exception as e:
    logger.critical(
        f"Critical Error: An unexpected error occurred while loading projection data: {e}", exc_info=True)
    print(
        f"[bold red]Критическая ошибка: Непредвиденная ошибка при загрузке данных проекций: {e}[/bold red]")
    raise SystemExit("Unexpected error loading projection data")


def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate the great circle distance between two points on earth (specified in decimal degrees)"""
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * \
        math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371  # Radius of earth in kilometers
    return c * r


def detect_coordinate_anomalies(coordinates, threshold_km=20):
    """
    Detect anomalous coordinates in a sequence by looking for points that are
    significantly further away from the majority of other points.

    Args:
        coordinates: List of tuples (name, lon, lat)
        threshold_km: Distance threshold in kilometers

    Returns:
        Tuple of (is_anomalous, reason, anomalous_points)
    """
    if len(coordinates) < 3:
        # Not enough points to detect anomalies
        return False, None, []

    distances = []
    anomalous_points = []

    # Calculate distance from each point to every other point
    for i, (name_i, lon_i, lat_i) in enumerate(coordinates):
        point_distances = []
        for j, (name_j, lon_j, lat_j) in enumerate(coordinates):
            if i != j:
                dist = haversine_distance(lat_i, lon_i, lat_j, lon_j)
                point_distances.append(dist)

        # Calculate the average distance to other points
        avg_distance = sum(point_distances) / len(point_distances)
        distances.append((i, avg_distance))

    # Find points that are much further away from others
    for idx, avg_dist in distances:
        # If a point's average distance to others is larger than the threshold
        if avg_dist > threshold_km:
            point_name, lon, lat = coordinates[idx]
            anomalous_points.append((idx, point_name, lon, lat))

    if anomalous_points:
        anomaly_details = ', '.join([f"{point_name} ({lat}, {lon})"
                                     for _, point_name, lon, lat in anomalous_points])
        reason = f"Обнаружены аномальные координаты, значительно удаленные от других"
        return True, reason, anomalous_points

    return False, None, []


def process_coordinates(input_string, transformer) -> Tuple[Optional[List[Tuple[str, float, float]]], Optional[str]]:
    """Processes a string with metric coordinates, transforming them and checking for validity."""
    logger.debug(f"-- Начало обработки МСК для строки: '{input_string[:70]}...' --")
    
    msk_regex = r'(\d+):\s*([-\d.]+)\s*м\.,\s*([-\d.]+)\s*м\.'
    logger.debug(f"1. Поиск координат МСК с помощью regex")
    coordinates = re.findall(msk_regex, input_string)
    
    if not coordinates:
        logger.debug("  - Координаты МСК не найдены. Возвращаем пустой результат.")
        return [], None

    logger.debug(f"2. Найдено {len(coordinates)} совпадений: {coordinates}")
    results = []
    for i, x_str, y_str in coordinates:
        logger.debug(f"\n-- Обработка совпадения {i}: x='{x_str}', y='{y_str}' --")
        try:
            x = float(x_str)
            y = float(y_str)
            logger.debug(f"  - Конвертировано в float: x={x}, y={y}")

            if x == 0 and y == 0:
                logger.debug("  - Нулевые координаты (0,0). Пропуск.")
                continue
            
            logger.debug("  - Трансформация в WGS84...")
            lon, lat = transformer.transform(y, x)
            logger.debug(f"    - Результат трансформации: lon={lon}, lat={lat}")

            if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                reason = f"Координаты МСК вне допустимого диапазона WGS84 (lat={lat}, lon={lon}) после трансформации."
                logger.warning(f"{reason} Исходные: x={x}, y={y}.")
                return None, reason

            rounded_lon, rounded_lat = round(lon, 6), round(lat, 6)
            point_name = f"точка {i}"
            results.append((point_name, rounded_lon, rounded_lat))
            logger.debug(f"  - Координаты валидны и добавлены в результат: Имя='{point_name}', Lon={rounded_lon}, Lat={rounded_lat}")

        except Exception as e:
            reason = f"Ошибка трансформации МСК координат: {e}. Исходные: x='{x_str}', y='{y_str}'."
            logger.error(reason)
            return None, reason

    logger.debug(f"\n3. Финальная проверка обработанных МСК...")
    if not results:
        logger.debug("  - После фильтрации (например, нулевых координат) не осталось валидных точек.")
        return [], None
    
    if len(results) >= 3:
        logger.debug("  - Запуск детектора аномалий для >= 3 точек.")
        is_anomalous, reason, _ = detect_coordinate_anomalies(results)
        if is_anomalous:
            logger.warning(f"  - Детектор аномалий сообщил: {reason}")
            return None, reason
        logger.debug("  - Аномалий не обнаружено.")

    logger.debug(f"4. Обработка МСК успешно завершена. Найдено {len(results)} валидных координат.")
    return results, None


def parse_coordinates(coord_str: str) -> Tuple[Optional[List[Tuple[str, float, float]]], Optional[str]]:
    """Парсит строку с координатами, проверяет их валидность и возвращает список кортежей (имя, долгота, широта) или ошибку.

    Returns:
        Tuple[Optional[List[Tuple[str, float, float]]], Optional[str]]:
            (список_координат, None) при успехе, (None, причина_ошибки) при ошибке.
            Список координат может быть пустым, если в строке не найдено валидных данных.
    """
    if not coord_str or not isinstance(coord_str, str):
        logger.debug("Пустая или нестроковая строка координат")
        return [], None

    coord_str = coord_str.strip()
    logger.debug(f"1. Исходная строка после удаления пробелов: '{coord_str}'")

    if not coord_str:
        logger.debug("Строка пуста после удаления пробелов. Возвращаем пустой результат.")
        return [], None

    logger.debug("2. Проверка типа координат...")

    if 'гск' in coord_str.lower():
        logger.debug("  - Обнаружен маркер 'гск'. Приоритет отдается парсингу ДМС (градусы, минуты, секунды).")
    else:
        logger.debug("  - Маркер 'гск' не найден.")
        if (' м.' in coord_str or ', м.' in coord_str or coord_str.endswith('м.')) and '°' not in coord_str:
            logger.debug("  - Обнаружен маркер 'м.' и отсутствует маркер '°'. Попытка парсинга как МСК (метровые).")
            for key, transformer in transformers.items():
                if key in coord_str:
                    logger.debug(f"    - Найдена известная система координат: '{key}'. Вызов process_coordinates.")
                    return process_coordinates(coord_str, transformer)
            reason = "Обнаружены координаты 'м.', но не найдена известная система координат МСК в строке."
            logger.warning(f"{reason} Строка: '{coord_str[:50]}'")
            return None, reason

    logger.debug("3. Проверка на наличие маркера ДМС ('°')...")
    if '°' not in coord_str:
        logger.debug("  - Маркер '°' не найден. Предполагается, что в строке нет координат. Возвращаем пустой результат.")
        return [], None

    logger.debug("  - Маркер '°' найден. Начинается парсинг ДМС.")
    parts = coord_str.split(';')
    logger.debug(f"4. Строка разделена на {len(parts)} частей по символу ';': {parts}")

    result = []
    has_valid_dms = False

    for i, part in enumerate([p.strip() for p in parts if p.strip()]):
        logger.debug(f"\n-- Обработка части {i+1}: '{part}' --")
        point_prefix = ""
        dms_regex = r'(\d+)°\s*(\d+)[\'′]\s*(\d+(?:[.,]\d+)?)[\"″′′]'
        logger.debug(f"  - Поиск ДМС с помощью regex: {dms_regex}")
        coords_match = re.findall(dms_regex, part)

        if not coords_match:
            logger.debug("  - Совпадений ДМС не найдено в этой части. Пропуск.")
            continue
        logger.debug(f"  - Найдено {len(coords_match)} совпадений ДМС: {coords_match}")

        if "выпуск" in part.lower():
            match = re.search(r'(выпуск\s+№?\s*\d+)', part, re.IGNORECASE)
            if match:
                point_prefix = match.group(1).strip()
                logger.debug(f"  - Извлечен префикс имени: '{point_prefix}'")
        elif "точка" in part.lower():
            match = re.search(r'(точка\s*\d+)', part, re.IGNORECASE)
            if match:
                point_prefix = match.group(1).strip()
                logger.debug(f"  - Извлечен префикс имени: '{point_prefix}'")

        if len(coords_match) % 2 != 0:
            reason = f"Нечетное количество найденных ДМС координат ({len(coords_match)}). Ожидается пара (широта, долгота)."
            logger.warning(reason)
            return None, reason

        if len(coords_match) >= 2:
            has_valid_dms = True
            logger.debug("  - Найдено достаточное количество совпадений для формирования пар координат.")
            for j in range(0, len(coords_match), 2):
                try:
                    lat_parts = coords_match[j]
                    lon_parts = coords_match[j+1]
                    logger.debug(f"    - Пара {j//2 + 1}: Широта (parts)={lat_parts}, Долгота (parts)={lon_parts}")

                    lat = sum(float(x.replace(',', '.')) / (60 ** k) for k, x in enumerate(lat_parts))
                    lon = sum(float(x.replace(',', '.')) / (60 ** k) for k, x in enumerate(lon_parts))
                    logger.debug(f"      - Конвертировано в десятичные: lat={lat}, lon={lon}")

                    if "ЮШ" in part or "S" in part:
                        lat = -lat
                        logger.debug("      - Обнаружен южный идентификатор (ЮШ/S). Широта инвертирована.")
                    if "ЗД" in part or "W" in part:
                        lon = -lon
                        logger.debug("      - Обнаружен западный идентификатор (ЗД/W). Долгота инвертирована.")

                    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                        reason = f"Координаты ДМС вне допустимого диапазона WGS84 (lat={lat}, lon={lon})."
                        logger.warning(reason)
                        return None, reason

                    point_specific_name = point_prefix
                    if len(coords_match) > 2 and point_prefix:
                        point_specific_name = f"{point_prefix}.{j // 2 + 1}"
                    elif len(coords_match) > 2 and not point_prefix:
                        point_specific_name = f"т.{j // 2 + 1}"
                    logger.debug(f"      - Итоговое имя точки: '{point_specific_name.strip()}'")

                    if lat != 0 or lon != 0:
                        result.append((point_specific_name.strip(), round(lon, 6), round(lat, 6)))
                        logger.debug("      - Координаты не нулевые и добавлены в результат.")
                    else:
                        logger.debug("      - Координаты нулевые и пропущены.")
                except Exception as e:
                    reason = f"Внутренняя ошибка при обработке пары ДМС: {e}."
                    logger.error(reason)
                    return None, reason

    logger.debug("\n5. Финальная проверка...")
    if '°' in coord_str and not has_valid_dms:
        reason = "Обнаружен маркер '°', но не найдено валидных пар ДМС координат."
        logger.warning(f"{reason} Строка: '{coord_str[:50]}'")
        return None, reason

    if result and len(result) >= 3:
        logger.debug("  - Запуск детектора аномалий для >= 3 точек.")
        is_anomalous, reason, _ = detect_coordinate_anomalies(result)
        if is_anomalous:
            logger.warning(f"  - Детектор аномалий сообщил: {reason}")
            return None, reason
        logger.debug("  - Аномалий не обнаружено.")

    logger.debug(f"6. Парсинг успешно завершен. Найдено {len(result)} валидных координат.")
    return result, None

def find_column_index(sheet, target_names: List[str], exact_match: bool = False) -> int:
    """Находит индекс столбца для любого из заданных имен заголовков в строках 1-5.

    Args:
        sheet: Лист Excel для поиска.
        target_names: Список имен заголовков для поиска.
        exact_match: Если True, требуется точное совпадение заголовка, иначе ищет подстроку (по умолчанию False).

    Returns:
        Индекс столбца или -1, если не найдено.
    """
    target_names_lower = [name.lower() for name in target_names]
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for idx, cell in enumerate(row):
            if cell:
                cell_str_lower = str(cell).lower()
                for target_name_lower in target_names_lower:
                    if (exact_match and cell_str_lower == target_name_lower) or \
                       (not exact_match and target_name_lower in cell_str_lower):
                        return idx
    return -1


def get_column_indices(sheet) -> dict:
    """Получает индексы всех необходимых столбцов."""
    columns = {
        "name": ["№ п/п"],
        "coord": ["Место водопользования"],
        "organ": ["Уполномоченный орган"],
        "additional_name": ["Наименование водного объекта"],
        "goal": ["Цель водопользования"],
        "vid": ["Вид водопользования"],
        "owner": ["Наименование"],
        "inn": ["ИНН"],
        "start_date": ["Дата начала водопользования"],
        "end_date": ["Дата окончания водопользования", "Дата прекращения действия"]
    }

    # Dictionary to specify which columns need exact matching
    exact_matches = {
        "owner": True
    }

    indices = {}
    for key, value in columns.items():
        # Use exact_match parameter if specified for this key
        exact = exact_matches.get(key, False)
        indices[key] = find_column_index(sheet, value, exact_match=exact)

    # Use the first name from the list for reporting missing columns
    original_names = {key: value[0] for key, value in columns.items()}
    for key, value in indices.items():
        if value == -1:
            logger.debug(
                f"Столбец '{original_names[key]}' (или его альтернативы) не найден.")

    return indices


def create_kml_point(kml, name: str, coords: Tuple[float, float], description: str, color: str) -> None:
    """Создает точку KML с заданными параметрами."""
    point = kml.newpoint(name=name, coords=[coords])
    point.description = description
    point.style.iconstyle.color = color
    point.style.iconstyle.scale = 1.0
    point.style.labelstyle.scale = 0.8


def create_kml_from_coordinates(sheet, output_file: str = "output.kml", sort_numbers: Optional[List[int]] = None) -> bool:
    """Создает KML-файл из листа с координатами и сохраняет аномалии в отдельный файл. Returns True on success, False otherwise."""
    kml = simplekml.Kml()
    indices = get_column_indices(sheet)
    anomalies_list = []  # Initialize list to store anomalies

    # Default min_row value
    min_row = 5

    # Check rows 2..5 using iter_rows()
    for row in sheet.iter_rows(min_row=2, max_row=5):
        cell = row[indices["coord"]] if indices["coord"] != -1 else None
        value = cell.value

        if isinstance(value, str) and ('м.' in value or '"' in value):
            min_row = cell.row  # Get the actual row number (3 or 4)
            break  # Stop at first match

    # Используем определенное значение min_row в цикле
    for row_idx, row in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), start=min_row):
        coords_str = row[indices["coord"]] if indices["coord"] != -1 else None
        if not isinstance(coords_str, str):
            continue
        main_name = row[indices["name"]
                        ] if indices["name"] != -1 else f"Row {row_idx}"
        logger.info(f"------------")

        # Вызываем обновленную функцию парсинга
        coords_array, error_reason = parse_coordinates(coords_str)

        # Если parse_coordinates вернула ошибку, пропускаем строку (она будет обработана как аномальная в другом модуле)
        if error_reason is not None:
            # Логирование уже произошло внутри parse_coordinates
            logger.warning(
                f"Строка {row_idx} (№ п/п {main_name}) пропущена из-за ошибки парсинга: {error_reason}")
            # Add anomaly details to the list
            anomalies_list.append({
                "row_index": row_idx,
                "main_name": main_name,
                "reason": error_reason,
                "coords_str": coords_str
            })
            continue  # Переходим к следующей строке Excel

        # Если coords_array это пустой список, значит парсинг прошел успешно, но валидных координат не найдено.
        # В этом случае просто пропускаем создание KML геометрии для этой строки.
        if not coords_array:
            logger.debug(
                f"Строка {row_idx} (№ п/п {main_name}) не содержит валидных координат для KML.")
            continue  # Нет точек для добавления в KML

        logger.info(
            f"Строка {row_idx} (№ п/п {main_name}): Распознано {len(coords_array)} точек.")

        if coords_array:  # Убедимся еще раз, что список не пуст
            color = generate_random_color()

            # Подготавливаем описание
            desc = []
            for key, column_name in [
                ("organ", "Уполномоченный орган"),
                ("additional_name", "Наименование водного объекта"),
                ("goal", "Цель водопользования"),
                ("vid", "Вид водопользования"),
                ("owner", "Владелец"),
                ("inn", "ИНН"),
                ("start_date", "Дата начала водопользования"),
                ("end_date", "Дата окончания водопользования"),
                ("coord", "Место водопользования")
            ]:
                if indices[key] != -1 and row[indices[key]]:
                    # Форматируем даты без времени, если это даты начала или окончания водопользования
                    if key in ["start_date", "end_date"] and hasattr(row[indices[key]], "date"):
                        # Если это объект datetime, берем только дату
                        date_value = row[indices[key]].date()
                        desc.append(f"{column_name}: {date_value}")
                    elif key in ["start_date", "end_date"] and isinstance(row[indices[key]], str):
                        # Если это строка, обрезаем время, если оно есть
                        date_str = row[indices[key]].split(" ")[0]
                        desc.append(f"{column_name}: {date_str}")
                    else:
                        desc.append(f"{column_name}: {row[indices[key]]}")
            
            description = '\n'.join(desc)
            description += "\n == Разработано RUDI.ru =="

            # Проверяем, есть ли 16-й столбец

            # Проверяем, есть ли более 3 точек и 16-й столбец не равен нулю или пуст
            skip_terms = ["Сброс сточных", "Забор (изъятие)"]
            if len(coords_array) > 3 and not any(term in row[indices["goal"]] for term in skip_terms):
                logger.debug(
                    f"Строка {row_idx} (№ п/п {main_name}): Создание полигона")
                # Создаем полигон
                polygon = kml.newpolygon(name=f"№ п/п {main_name}")

                # Сортируем координаты только если main_name есть в sort_numbers
                if (sort_numbers and int(main_name) in sort_numbers) or len(coords_array) == 4:
                    sorted_coords = sort_coordinates(
                        [(lon, lat) for _, lon, lat in coords_array])
                else:
                    sorted_coords = [(lon, lat)
                                     for _, lon, lat in coords_array]

                polygon.outerboundaryis = sorted_coords # type: ignore
                polygon.style.linestyle.color = color
                polygon.style.linestyle.width = 3
                polygon.style.polystyle.color = simplekml.Color.changealphaint(
                    100, color)
                polygon.description = description
            else:
                # Создаем линию, если есть несколько точек и условия выполнены
                if len(coords_array) > 2 \
                        and all(name.startswith("точка") for name, _, _ in coords_array) \
                        and row[indices["goal"]] != "Сброс сточных вод":
                    logger.debug(
                        f"Строка {row_idx} (№ п/п {main_name}): Создание линии")
                    line = kml.newlinestring(name=f"№ п/п {main_name}",
                                             coords=[(lon, lat) for _, lon, lat in coords_array])
                    line.style.linestyle.color = color
                    line.style.linestyle.width = 3
                    line.description = description
                else:
                    # Создаем отдельные точки, если линия не была создана
                    index = 1
                    for point_name, lon, lat in coords_array:
                        logger.debug(f"  Точка: {point_name} ({lat}, {lon})")
                        # print(f"{lat}, {lon}")
                        if row[indices["goal"]] == "Сброс сточных вод":
                            full_name = f"№ п/п {main_name} - сброс {index}"
                        else:
                            full_name = f"№ п/п {main_name} - {point_name}" if point_name else f"№ п/п {main_name}"
                        create_kml_point(
                            kml, full_name, (lon, lat), description, color)
                        index += 1

    kml.save(output_file)

    anomaly_file_created = False  # Initialize return value
    if anomalies_list and output_file:
        # Use current dir if output_file has no path
        output_dir = os.path.dirname(output_file) or '.'
        original_basename = os.path.basename(output_file)
        # Capture the return value from save_anomalies_to_excel
        anomaly_file_created = save_anomalies_to_excel(
            anomalies_list, original_basename, output_dir)
    elif anomalies_list and not output_file:
        logger.warning(
            "Anomalies were detected, but the original filename was not provided. Anomalies will not be saved to a separate file.")

    return anomaly_file_created  # Return the status


def save_anomalies_to_excel(anomalies: List[dict], original_basename: str, output_directory: str) -> bool:
    """Saves detected anomalies to a separate Excel file in the specified output directory. Returns True on success, False otherwise."""
    if not anomalies:
        return False  # Nothing to save

    # Construct the output filename using the original basename
    name, ext = os.path.splitext(original_basename)
    output_filename = f"ANO_{name}.xlsx"
    # Place it in the output directory
    output_path = os.path.join(output_directory, output_filename)

    logger.info(f"Saving {len(anomalies)} anomalies to '{output_path}'...")

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    # Write headers
    headers = ["Строка в оригинальном файле", "№ п/п", "Причина", "Координаты"]
    ws.append(headers)

    # Write anomaly data
    for anomaly in anomalies:
        ws.append([
            anomaly.get("row_index", "N/A"),
            anomaly.get("main_name", "N/A"),
            anomaly.get("reason", "N/A"),
            anomaly.get("coords_str", "N/A"),
        ])

    # Adjust column widths (optional, for better readability)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            # Check if cell.value is not None and convert to string
            if cell.value is not None:
                try:
                    value_str = str(cell.value)
                    if len(value_str) > max_length:
                        max_length = len(value_str)
                except Exception as e:
                    logger.warning(
                        f"Could not determine length for cell value {cell.value} in column {column}: {e}")
        adjusted_width = (max_length + 2)
        if adjusted_width > 64:
            adjusted_width = 64
        ws.column_dimensions[column].width = adjusted_width

    try:
        wb.save(output_path)
        logger.info(f"Anomalies successfully saved to '{output_path}'.")
        return True  # Return True on successful save
    except Exception as e:
        logger.error(
            f"Failed to save anomalies to '{output_path}': {e}", exc_info=True)
        print(
            f"[bold red]Ошибка при сохранении файла аномалий '{output_path}': {e}[/bold red]")
        return False  # Return False on error

if __name__ == "__main__":
    # Ensure the logger is at DEBUG level for the console when in debug mode
    # This is a bit of a hack, but it ensures debug messages are seen without changing the global config
    for handler in logger.root.handlers:
        if isinstance(handler, logging.StreamHandler):
            handler.setLevel(logging.DEBUG)

    print("--- Режим отладки парсера координат ---")
    print("Введите строку для парсинга. Для выхода введите 'exit' или 'quit'.")

    while True:
        input_string = input("> ")
        if input_string.lower() in ["exit", "quit"]:
            break

        if not input_string:
            continue

        logger.info(f"--- Начало парсинга строки: '{input_string}' ---")
        coords, reason = parse_coordinates(input_string)
        print("\n--- Итоговый результат ---")

        if reason:
            print(f"Ошибка: {reason}")
        elif not coords:
            print("Координаты не найдены или являются нулевыми.")
        else:
            print(f"Успешно найдено {len(coords)} координат:")
            for i, (name, lon, lat) in enumerate(coords):
                print(f"  {i+1}. Имя: '{name}', Долгота: {lon}, Широта: {lat}")
        print("--------------------------\n")
