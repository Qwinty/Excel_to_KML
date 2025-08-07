import glob
import os
import logging
import time
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any
from collections import defaultdict

from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.prompt import Prompt, Confirm, IntPrompt
from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn, SpinnerColumn
from rich.text import Text
from rich import traceback

from xlsx_to_kml import create_kml_from_coordinates, parse_coordinates, process_coordinates, transformers, create_transformer, ConversionResult
from separator import split_excel_file_by_merges
from utils import setup_logging

# Set up rich traceback for better error display
traceback.install(show_locals=True)

# Console setup (logging will be set up in main())
console = Console()

# Logger will be initialized in main() but we need a module-level reference
logger = None

# --- Configuration ---


@dataclass
class Config:
    """Application configuration."""
    input_dir: str = "input"
    xlsx_output_dir: str = "output/xlsx"
    kml_output_dir: str = "output/kml"
    single_kml_output_dir: str = "output/kml_single"
    header_rows_count: int = 5
    merge_columns: tuple = (1, 7)  # Columns A-G
    # None = auto-detect based on CPU count
    max_parallel_workers: Optional[int] = None


# Config will be created in main() and passed to functions that need it

# --- Worker Process Initialization ---

def initialize_worker_logging():
    """Initializer for each worker process to set up its logging.
    
    This function is called once when each worker process starts up.
    It configures logging so that:
    - Console shows only ERROR and CRITICAL messages
    - Files still log everything from DEBUG up
    """
    setup_logging(console_level=logging.ERROR)

# --- Statistics Data Structures ---
# ConversionResult is now imported from xlsx_to_kml


@dataclass
class ProcessingStats:
    """Aggregate statistics for the entire processing session."""
    start_time: float = field(default_factory=time.time)
    regions_detected: int = 0
    files_created: List[str] = field(default_factory=list)
    file_results: Dict[str, ConversionResult] = field(default_factory=dict)
    conversion_errors: int = 0
    anomaly_files_generated: int = 0

    def add_file_result(self, result: ConversionResult):
        """Add a file conversion result to the statistics."""
        self.file_results[result.filename] = result

    def get_processing_time(self) -> float:
        """Get total processing time in seconds."""
        return time.time() - self.start_time

    def get_total_stats(self) -> Dict[str, int]:
        """Calculate aggregate statistics across all files."""
        totals = {
            'total_files': len(self.file_results),
            'total_rows': 0,
            'successful_rows': 0,
            'failed_rows': 0,
            'anomaly_rows': 0
        }

        for result in self.file_results.values():
            totals['total_rows'] += result.total_rows
            totals['successful_rows'] += result.successful_rows
            totals['failed_rows'] += result.failed_rows
            totals['anomaly_rows'] += result.anomaly_rows

        return totals

    def get_most_problematic_files(self, top_n: int = 5) -> List[ConversionResult]:
        """Get the most problematic files sorted by failure rate."""
        files_with_issues = [
            result for result in self.file_results.values()
            if result.total_rows > 0 and result.failure_rate > 0
        ]

        # Sort by failure rate (highest first)
        sorted_files = sorted(
            files_with_issues, key=lambda x: x.failure_rate, reverse=True)
        return sorted_files[:top_n]

    def calculate_quality_score(self) -> Dict[str, Any]:
        """Calculate data quality scores."""
        totals = self.get_total_stats()

        if totals['total_rows'] == 0:
            return {'parsing': 0, 'completeness': 0, 'consistency': 0, 'overall': 0}

        # Coordinate parsing score (0-100)
        parsing_score = (totals['successful_rows'] /
                         totals['total_rows']) * 100

        # Data completeness score (simplified - based on rows with any data)
        # For now, assume any processed row has some completeness
        completeness_score = max(
            0, 100 - (totals['failed_rows'] / totals['total_rows']) * 50)

        # Format consistency score (based on error variety)
        all_errors = []
        for result in self.file_results.values():
            all_errors.extend(result.error_reasons)

        # Simple consistency metric: fewer unique error types = more consistent
        unique_errors = len(set(all_errors)) if all_errors else 0

        # Analyze error types for user-friendly display
        error_analysis = None
        if all_errors:
            from collections import Counter
            import re

            # Group similar errors by pattern
            grouped_errors = Counter()
            error_patterns = {
                r'Нечетное количество найденных ДМС координат \(\d+\)': 'Нечетное количество найденных ДМС координат',
                r'Нечетное количество найденных ЛМС координат \(\d+\)': 'Нечетное количество найденных ЛМС координат',
                r'Координаты ДМС вне допустимого диапазона WGS84 \(lat=[-\d.]+, lon=[-\d.]+\)': 'Координаты ДМС вне допустимого диапазона WGS84',
                r'Координаты МСК вне допустимого диапазона WGS84 \(lat=[-\d.]+, lon=[-\d.]+\)': 'Координаты МСК вне допустимого диапазона WGS84',
                r'Ошибка трансформации МСК координат: .+': 'Ошибка трансформации МСК координат',
                r'Обнаружены аномальные координаты, значительно удаленные от других': 'Обнаружены аномальные координаты, значительно удаленные от других'
            }

            for error in all_errors:
                grouped = False
                for pattern, group_name in error_patterns.items():
                    if re.match(pattern, error):
                        grouped_errors[group_name] += 1
                        grouped = True
                        break
                if not grouped:
                    # If no pattern matches, use the original error (truncated)
                    display_error = error[:80] + \
                        "..." if len(error) > 80 else error
                    grouped_errors[display_error] += 1

            error_analysis = {
                'total_errors': len(all_errors),
                'unique_types': len(grouped_errors),
                # Top 10 most frequent error groups
                'top_errors': grouped_errors.most_common(10)
            }

        # More reasonable penalty: max penalty should be around 50%, not 100%
        # This way, even with many error types, we don't go to 0%
        # 2% penalty per unique error type, minimum 20%
        consistency_score = max(20, 100 - (unique_errors * 2))

        # Overall weighted score
        overall = (parsing_score * 0.5 + completeness_score *
                   0.3 + consistency_score * 0.2)

        return {
            'parsing': round(parsing_score, 1),
            'completeness': round(completeness_score, 1),
            'consistency': round(consistency_score, 1),
            'overall': round(overall, 1),
            'error_analysis': error_analysis
        }


def process_file_worker(
    xlsx_file_path: str,
    kml_file_path: str,
    xlsx_output_dir: str,
    kml_output_dir: str
) -> Tuple[bool, str, Optional[ConversionResult], Optional[str]]:
    """
    Worker function for parallel file processing.

    Args:
        xlsx_file_path: Path to the source Excel file
        kml_file_path: Path for the output KML file
        xlsx_output_dir: Directory containing Excel output files
        kml_output_dir: Directory for KML output files

    Returns:
        Tuple of (success, filename, conversion_result, error_message)
    """

    try:
        # Extract filename for logging
        filename = Path(xlsx_file_path).name

        # Ensure the target directory for the KML file exists
        Path(kml_file_path).parent.mkdir(parents=True, exist_ok=True)

        # Load workbook (ensure data_only=True)
        workbook = load_workbook(
            filename=xlsx_file_path, data_only=True, read_only=True)

        # Perform KML conversion
        conversion_result = create_kml_from_coordinates(
            workbook.active,
            output_file=kml_file_path,
            filename=filename
        )

        return True, filename, conversion_result, None

    except Exception as e:
        filename = Path(xlsx_file_path).name if xlsx_file_path else "Unknown"
        error_message = f"Error converting {filename}: {str(e)}"
        return False, filename, None, error_message


def display_error_analysis(error_analysis: Dict[str, Any]):
    """Display error analysis in a user-friendly table format."""
    if not error_analysis or not error_analysis.get('top_errors'):
        return

    error_table = Table(show_header=True, header_style="bold yellow")
    error_table.add_column("№", width=3, justify="center")
    error_table.add_column("Тип ошибки", min_width=40)
    error_table.add_column("Количество", justify="right", style="red")
    error_table.add_column("Процент", justify="right", style="bright_yellow")

    total_errors = error_analysis['total_errors']

    for i, (error_type, count) in enumerate(error_analysis['top_errors'], 1):
        percentage = (count / total_errors) * 100
        # Truncate very long error messages for better display
        display_error = error_type[:60] + \
            "..." if len(error_type) > 60 else error_type
        error_table.add_row(
            str(i),
            display_error,
            str(count),
            f"{percentage:.1f}%"
        )

    # Summary row
    if len(error_analysis['top_errors']) < error_analysis['unique_types']:
        remaining_types = error_analysis['unique_types'] - \
            len(error_analysis['top_errors'])
        remaining_count = total_errors - \
            sum(count for _, count in error_analysis['top_errors'])
        remaining_percentage = (
            remaining_count / total_errors) * 100 if total_errors > 0 else 0

        error_table.add_row(
            "...",
            f"Другие типы ошибок ({remaining_types} типов)",
            str(remaining_count),
            f"{remaining_percentage:.1f}%",
            style="dim"
        )

    console.print(Panel(
        error_table,
        title=f"🔍 Анализ ошибок ({error_analysis['unique_types']} уникальных типов, {total_errors} всего)",
        border_style="yellow"
    ))


def _format_processing_time(processing_time: float) -> str:
    """Format processing time into human-readable string."""
    if processing_time < 60:
        return f"{processing_time:.1f}с"
    else:
        minutes = int(processing_time // 60)
        seconds = int(processing_time % 60)
        return f"{minutes}м {seconds}с"


def _display_processing_summary(stats: ProcessingStats, totals: Dict[str, int], time_str: str):
    """Display processing summary statistics."""
    success_rate = (totals['successful_rows'] / totals['total_rows']
                    * 100) if totals['total_rows'] > 0 else 0

    summary_table = Table(show_header=False, box=None, padding=(0, 1))
    summary_table.add_column("Параметр", style="bold", width=25)
    summary_table.add_column("Значение", style="green")

    summary_table.add_row("Файлов обнаружено:",
                          f"{stats.regions_detected} регионов")
    if stats.anomaly_files_generated > 0:
        summary_table.add_row("Файлы с аномалиями:",
                              f"{stats.anomaly_files_generated} файла")
    summary_table.add_row("Объектов обработано:",
                          f"{totals['total_rows']} строк → {totals['successful_rows']} успешно ({success_rate:.1f}%)")
    summary_table.add_row("Время обработки:", time_str)

    console.print(Panel(
        summary_table,
        title="📊 Сводка обработки",
        border_style="cyan"
    ))


def _display_problematic_files(stats: ProcessingStats):
    """Display most problematic files if there are any issues."""
    problematic_files = stats.get_most_problematic_files(7)
    if not problematic_files:
        return

    problem_table = Table(show_header=True, header_style="bold red")
    problem_table.add_column("№", width=3, justify="center")
    problem_table.add_column("Файл", min_width=30)
    problem_table.add_column("Проблемные строки", justify="right", style="red")
    problem_table.add_column("Процент", justify="right", style="yellow")

    for i, result in enumerate(problematic_files, 1):
        problem_table.add_row(
            str(i),
            result.filename,
            f"{result.failed_rows}/{result.total_rows} объектов",
            f"{result.failure_rate:.1f}%"
        )

    console.print(Panel(
        problem_table,
        title=f"⚠️ Наиболее проблемные файлы (топ {len(problematic_files)})",
        border_style="red"
    ))


def _create_progress_bar(value: float, width: int = 20) -> str:
    """Create a visual progress bar for quality metrics."""
    filled = int(value / 5)  # Each block represents 5%
    empty = width - filled
    return "█" * filled + "▌" * (1 if value % 5 >= 2.5 else 0) + "░" * (empty - (1 if value % 5 >= 2.5 else 0))


def _get_quality_grade_and_color(overall_score: float) -> Tuple[str, str]:
    """Get quality grade and color based on overall score."""
    if overall_score >= 90:
        return "green", "Отлично"
    elif overall_score >= 80:
        return "bright_green", "Хорошо"
    elif overall_score >= 70:
        return "yellow", "Удовлетворительно"
    elif overall_score >= 60:
        return "bright_red", "Плохо"
    else:
        return "red", "Очень плохо"


def _display_quality_scores(quality_scores: Dict[str, Any]):
    """Display data quality scores with visual progress bars."""
    overall_score = quality_scores['overall']
    overall_color, overall_grade = _get_quality_grade_and_color(overall_score)

    console.print(Panel(
        f"[bold {overall_color}]Общая оценка качества: {overall_score:.0f}/100 ({overall_grade})[/bold {overall_color}]\n\n"
        f"• Парсинг координат: {quality_scores['parsing']:.1f}% {_create_progress_bar(quality_scores['parsing'])} ({quality_scores['parsing']:.0f}/100)\n"
        f"• Полнота данных: {quality_scores['completeness']:.1f}% {_create_progress_bar(quality_scores['completeness'])} ({quality_scores['completeness']:.0f}/100)\n"
        f"• Согласованность форматов: {quality_scores['consistency']:.1f}% {_create_progress_bar(quality_scores['consistency'])} ({quality_scores['consistency']:.0f}/100)",
        title="🎯 Оценка качества данных",
        border_style="blue"
    ))


def _display_additional_info(stats: ProcessingStats):
    """Display additional information about conversion errors."""
    if stats.conversion_errors > 0:
        console.print(
            f"[yellow]⚠️ Дополнительно: {stats.conversion_errors} файлов не удалось обработать из-за критических ошибок.[/yellow]")


def display_processing_statistics(stats: ProcessingStats):
    """Display comprehensive processing statistics using Rich components."""
    if not stats.file_results:
        console.print(
            "[yellow]Нет данных для отображения статистики.[/yellow]")
        return

    totals = stats.get_total_stats()
    processing_time = stats.get_processing_time()
    quality_scores = stats.calculate_quality_score()
    time_str = _format_processing_time(processing_time)

    # Display all sections
    _display_processing_summary(stats, totals, time_str)
    _display_problematic_files(stats)
    _display_quality_scores(quality_scores)

    # Error analysis (if there are errors to analyze)
    if quality_scores.get('error_analysis'):
        display_error_analysis(quality_scores['error_analysis'])

    _display_additional_info(stats)
    console.print()


def choose_file(config: Config) -> Optional[str]:
    """Prompt user to choose an Excel file from the input directory using Rich interface."""
    input_dir = Path(config.input_dir)

    # Create input directory if it doesn't exist
    if not input_dir.exists():
        input_dir.mkdir(parents=True, exist_ok=True)
        console.print(Panel(
            f"[yellow]Создана папка '{input_dir}'. Пожалуйста, поместите Excel файлы в нее.[/yellow]",
            title="Информация",
            border_style="yellow"
        ))
        return None

    # Find Excel files
    files = list(input_dir.glob("*.xlsx"))

    if not files:
        console.print(Panel(
            f"[bold red]Файлы Excel (.xlsx) не найдены в папке '{input_dir}'.[/bold red]",
            title="Ошибка",
            border_style="red"
        ))
        return None

    # Create table with file information
    table = Table(title="Доступные файлы Excel",
                  show_header=True, header_style="bold cyan")
    table.add_column("№", style="dim", width=4, justify="center")
    table.add_column("Имя файла", min_width=20)
    table.add_column("Размер", justify="right", style="green")
    table.add_column("Дата изменения", justify="center", style="blue")

    for i, file_path in enumerate(files, 1):
        size_kb = file_path.stat().st_size / 1024
        size_str = f"{size_kb:.1f} KB" if size_kb < 1024 else f"{size_kb/1024:.1f} MB"
        mod_time = file_path.stat().st_mtime
        import datetime
        mod_date = datetime.datetime.fromtimestamp(
            mod_time).strftime("%Y-%m-%d %H:%M")

        table.add_row(
            str(i),
            file_path.name,
            size_str,
            mod_date
        )

    console.print(table)

    # Get user choice with validation
    try:
        choice = IntPrompt.ask(
            "Выберите номер файла",
            choices=[str(i) for i in range(1, len(files) + 1)],
            show_choices=False
        )
        return str(files[choice - 1])
    except (KeyboardInterrupt, EOFError):
        console.print("\n[yellow]Выбор отменен.[/yellow]")
        return None


def _setup_debug_logging() -> Tuple[Optional[logging.Handler], Optional[int]]:
    """Setup debug logging for coordinate parser mode."""
    root_logger = logging.getLogger()
    console_handler = None
    original_console_level = None

    for handler in root_logger.handlers:
        if isinstance(handler, logging.StreamHandler):
            console_handler = handler
            original_console_level = handler.level
            handler.setLevel(logging.DEBUG)
            logger.debug(
                "Установлен DEBUG уровень логирования для консоли в режиме отладки")
            break

    return console_handler, original_console_level


def _cleanup_debug_logging(console_handler: Optional[logging.Handler], original_console_level: Optional[int]):
    """Restore original logging level after debug mode."""
    if console_handler and original_console_level is not None:
        logger.debug(
            f"Восстановлен исходный уровень логирования консоли: {logging.getLevelName(original_console_level)}")
        console_handler.setLevel(original_console_level)


def _get_debug_mode_choice() -> str:
    """Display mode selection menu and get user choice."""
    mode_table = Table(title="Режимы парсинга", show_header=False, box=None)
    mode_table.add_column("№", style="bold cyan", width=3)
    mode_table.add_column("Описание", style="white")

    mode_table.add_row("1", "Автоматический режим (как в основной программе)")
    mode_table.add_row("2", "Ввести собственную proj4 строку")
    mode_table.add_row("3", "Вернуться в главное меню")

    console.print(mode_table)

    return Prompt.ask(
        "Введите номер режима",
        choices=["1", "2", "3"],
        show_choices=False
    )


def _get_custom_proj4_transformer() -> Tuple[Optional[Any], Optional[str]]:
    """Get custom proj4 transformer from user input."""
    console.print(Panel(
        "[bold cyan]Ввод собственной proj4 строки[/bold cyan]\n\n"
        "Введите proj4 строку для преобразования координат МСК.\n\n"
        "[dim]Пример:[/dim]\n"
        "[yellow]+proj=tmerc +lat_0=0 +lon_0=130.71666666667 +k=1 +x_0=4300000 +y_0=-16586.442 +ellps=krass +units=m +no_defs[/yellow]",
        title="Настройка proj4",
        border_style="cyan"
    ))

    while True:
        try:
            custom_proj4 = Prompt.ask(
                "\n[bold]Proj4 строка[/bold]",
                default="",
                show_default=False
            ).strip()

            if not custom_proj4:
                console.print("[yellow]Ввод не может быть пустым.[/yellow]")
                continue

            if custom_proj4.lower() in ["back", "назад"]:
                return None, None

            # Проверяем, что строка начинается с +proj
            if not custom_proj4.startswith('+proj'):
                console.print(
                    "[yellow]Proj4 строка должна начинаться с '+proj'.[/yellow]")
                continue

            # Пытаемся создать трансформер
            selected_transformer = create_transformer(custom_proj4)
            selected_proj4_name = "Пользовательская proj4"

            display_proj4 = custom_proj4[:60] + \
                "..." if len(custom_proj4) > 60 else custom_proj4
            console.print(
                f"[green]✓ Proj4 строка успешно загружена:[/green] [dim]{display_proj4}[/dim]")
            return selected_transformer, selected_proj4_name

        except (KeyboardInterrupt, EOFError):
            console.print("\n[yellow]Ввод отменен.[/yellow]")
            return None, None
        except Exception as e:
            console.print(Panel(
                f"[bold red]Ошибка при загрузке proj4 строки:[/bold red]\n{e}\n\n"
                "[yellow]Попробуйте ввести корректную proj4 строку или введите 'back' для возврата.[/yellow]",
                title="Ошибка",
                border_style="red"
            ))
            continue


def _parse_coordinate_string(input_string: str, mode_choice: str, selected_transformer: Optional[Any]) -> Tuple[Optional[List], Optional[str]]:
    """Parse coordinate string based on mode and transformer."""
    logger.info(f"--- Начало парсинга строки: '{input_string}' ---")

    if mode_choice == "1":
        # Автоматический режим
        return parse_coordinates(input_string)
    elif mode_choice == "2":
        # Ручной режим с пользовательской proj4 системой
        if (' м.' in input_string or ', м.' in input_string or input_string.endswith('м.')) and '°' not in input_string:
            return process_coordinates(input_string, selected_transformer)
        else:
            return parse_coordinates(input_string)

    return None, "Неизвестный режим парсинга"


def _display_parsing_results(coords: Optional[List], reason: Optional[str]):
    """Display parsing results in formatted panels."""
    if reason:
        console.print(Panel(
            f"[bold red]Ошибка:[/bold red] {reason}",
            title="❌ Результат парсинга",
            border_style="red"
        ))
    elif not coords:
        console.print(Panel(
            "[yellow]Координаты не найдены или являются нулевыми.[/yellow]",
            title="⚠️ Результат парсинга",
            border_style="yellow"
        ))
    else:
        # Создаем таблицу с результатами
        result_table = Table(
            title=f"✅ Найдено {len(coords)} координат", show_header=True, header_style="bold green")
        result_table.add_column("№", style="dim", width=3, justify="center")
        result_table.add_column("Имя", style="cyan")
        result_table.add_column("Долгота", style="green", justify="right")
        result_table.add_column("Широта", style="green", justify="right")

        for i, (name, lon, lat) in enumerate(coords, 1):
            result_table.add_row(str(i), name, f"{lon:.6f}", f"{lat:.6f}")

        console.print(result_table)

    console.print()  # Add spacing


def _run_coordinate_parsing_loop(mode_choice: str, selected_transformer: Optional[Any], selected_proj4_name: Optional[str]):
    """Run the main coordinate parsing input loop."""
    mode_text = 'Автоматический' if mode_choice == '1' else f'Ручной ({selected_proj4_name})'

    console.print(Panel(
        f"[bold green]Режим парсинга: {mode_text}[/bold green]\n\n"
        "Введите строку для парсинга координат.\n"
        "[dim]Для возврата к выбору режима введите 'back' или 'назад'.[/dim]",
        title="🔍 Парсинг координат",
        border_style="green"
    ))

    while True:
        try:
            input_string = Prompt.ask(
                "[bold cyan]Строка для парсинга[/bold cyan]")

            if input_string.lower() in ["back", "назад"]:
                break

            if not input_string.strip():
                continue

            coords, reason = _parse_coordinate_string(
                input_string, mode_choice, selected_transformer)
            _display_parsing_results(coords, reason)

        except (KeyboardInterrupt, EOFError):
            console.print("\n[yellow]Ввод отменен.[/yellow]")
            break


def debug_coordinate_parser():
    """Интерактивный отладочный парсер координат с выбором proj4 системы."""
    console.print(Panel(
        "[bold magenta]Режим отладки парсера координат[/bold magenta]\n"
        "[dim]Введите строки для парсинга координат и тестирования различных proj4 систем[/dim]",
        title="🔧 Отладка",
        border_style="magenta"
    ))

    console_handler, original_console_level = _setup_debug_logging()

    try:
        while True:
            mode_choice = _get_debug_mode_choice()

            if mode_choice == "3":
                break

            # Выбор proj4 системы для режима 2
            selected_transformer = None
            selected_proj4_name = None

            if mode_choice == "2":
                selected_transformer, selected_proj4_name = _get_custom_proj4_transformer()
                if not selected_transformer:
                    continue

            _run_coordinate_parsing_loop(
                mode_choice, selected_transformer, selected_proj4_name)

    finally:
        _cleanup_debug_logging(console_handler, original_console_level)


def display_welcome():
    """Display welcome screen with application info."""
    console.print(Panel.fit(
        "[bold magenta]Конвертер Excel в KML[/bold magenta]\n\n"
        "[dim]Преобразование файлов Excel с координатами в формат KML\n"
        "для использования в картографических приложениях[/dim]",
        title="🗺️ Excel to KML Converter (RUDI.RU)",
        border_style="bright_blue",
        padding=(1, 2)
    ))


def process_mode_1_full_processing(config: Config):
    """Обработка режима 1: Полная обработка с разделением и преобразованием в KML."""
    console.print(Panel(
        "[bold cyan]Режим: Разделение файла и преобразование в KML[/bold cyan]\n\n"
        "[dim]Этот режим выполнит полный цикл обработки:\n"
        "1. Разделение файла по регионам\n"
        "2. Преобразование каждого региона в KML[/dim]",
        title="🔄 Полная обработка",
        border_style="cyan"
    ))

    input_file = choose_file(config)
    if not input_file:
        return

    input_filename = Path(input_file).name

    # Initialize statistics collection
    processing_stats = ProcessingStats()

    # Display processing info
    info_table = Table(show_header=False, box=None)
    info_table.add_column("Параметр", style="bold", width=30)
    info_table.add_column("Значение", style="green")

    info_table.add_row("Входной файл:", input_filename)
    info_table.add_row("Выход (XLSX):", config.xlsx_output_dir)
    info_table.add_row("Выход (KML):", config.kml_output_dir)

    console.print(Panel(
        info_table,
        title="ℹ️ Параметры обработки",
        border_style="blue"
    ))

    # --- Stage 1: Separation ---
    separation_success = _process_file_separation(
        input_file, input_filename, processing_stats, config)

    # --- Stage 2: KML Conversion (only if separation was successful) ---
    if separation_success:
        _process_kml_conversion(processing_stats, config)

    # Display comprehensive statistics
    if separation_success:
        display_processing_statistics(processing_stats)


def _process_file_separation(input_file: str, input_filename: str, processing_stats: ProcessingStats, config: Config) -> bool:
    """Обработка этапа разделения файла."""
    separation_success = False

    console.print("[cyan]🔄 Этап 1: Разделение файла по регионам...[/cyan]")

    try:
        # Ensure the separated XLSX output directory exists
        Path(config.xlsx_output_dir).mkdir(parents=True, exist_ok=True)
        logger.info(
            f"Создана папка для разделенных XLSX: {config.xlsx_output_dir}")

        split_excel_file_by_merges(
            input_path=input_file,
            output_base_dir=config.xlsx_output_dir,
            header_rows_count=config.header_rows_count,
            merge_cols=config.merge_columns
        )

        # Count regions created
        separated_files = list(Path(config.xlsx_output_dir).rglob('*.xlsx'))
        processing_stats.regions_detected = len(separated_files)
        processing_stats.files_created = [str(f) for f in separated_files]

        separation_success = True

    except Exception as e:
        console.print(Panel(
            f"[bold red]Ошибка на этапе разделения:[/bold red]\n{e}\n\n"
            "[dim]Проверьте, что файл не открыт в Excel и доступен для чтения.[/dim]",
            title="❌ Ошибка этапа 1",
            border_style="red"
        ))
        logger.exception(
            f"Ошибка в режиме 1 (Разделение) при обработке файла {input_file}")

    if separation_success:
        console.print(Panel(
            f"[bold green]✅ Этап 1 завершен успешно[/bold green]\n\n"
            f"Файл '[cyan]{input_filename}[/cyan]' успешно разделен.\n"
            f"Разделенные XLSX файлы: [blue]{config.xlsx_output_dir}[/blue]",
            title="🎉 Разделение завершено",
            border_style="green"
        ))

    return separation_success


def _process_kml_conversion(processing_stats: ProcessingStats, config: Config):
    """Обработка этапа преобразования в KML."""
    console.print(Panel(
        "[bold cyan]Этап 2: Преобразование разделенных файлов в KML[/bold cyan]\n\n"
        "[dim]Поиск разделенных файлов и преобразование в формат KML...[/dim]",
        title="🔄 Этап 2",
        border_style="cyan"
    ))

    # Find all .xlsx files created by the separator
    separated_files = list(Path(config.xlsx_output_dir).rglob('*.xlsx'))

    if not separated_files:
        console.print(Panel(
            f"[yellow]Не найдено файлов *.xlsx для преобразования в KML в директории '{config.xlsx_output_dir}' и ее подпапках.[/yellow]",
            title="⚠️ Предупреждение",
            border_style="yellow"
        ))
        return

    console.print(
        f"[green]✓ Найдено {len(separated_files)} файлов .xlsx для преобразования.[/green]")

    # Ensure the KML output base directory exists
    Path(config.kml_output_dir).mkdir(parents=True, exist_ok=True)
    logger.info(f"Создана базовая папка для KML: {config.kml_output_dir}")

    conversion_errors = 0

    # Logging suppression is now handled by worker process initialization
    conversion_errors = _run_parallel_conversion(
        separated_files, processing_stats, config)

    # Reporting results
    _report_conversion_results(separated_files, conversion_errors, config)


def _run_parallel_conversion(separated_files: List[Path], processing_stats: ProcessingStats, config: Config) -> int:
    """Запуск параллельного преобразования файлов."""
    conversion_errors = 0

    # --- Use improved Progress with Rich ---
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total} файлов)"),
        TimeRemainingColumn(),
        console=console,
        transient=False
    ) as progress:
        # Add the conversion task
        task = progress.add_task(
            "Преобразование в KML...", total=len(separated_files))

        # Prepare arguments for parallel processing
        worker_args = _prepare_worker_args(separated_files, config)

        # Determine the number of workers based on configuration and CPU count
        max_workers = _determine_max_workers(separated_files, config)

        console.print(
            f"[dim]Запуск параллельной обработки с {max_workers} потоками...[/dim]")
        console.print(
            f"[dim]DEBUG/WARNING сообщения подавлены в консоли для повышения производительности[/dim]")

        # Process files in parallel
        with ProcessPoolExecutor(
            max_workers=max_workers,
            initializer=initialize_worker_logging
        ) as executor:
            # Submit all tasks
            future_to_file = {
                executor.submit(process_file_worker, **args): args['xlsx_file_path']
                for args in worker_args
            }

            # Process completed tasks as they finish
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                filename = Path(file_path).name

                try:
                    success, processed_filename, conversion_result, error_message = future.result()

                    # Print filename on separate line
                    if success:
                        console.print(
                            f"[dim]Завершено: [green]{processed_filename}[/green][/dim]")

                        # Add result to processing statistics
                        if conversion_result is not None:
                            processing_stats.add_file_result(conversion_result)

                            # Count anomaly files (check if anomaly file was created)
                            if conversion_result.anomaly_file_created:
                                processing_stats.anomaly_files_generated += 1
                    else:
                        console.print(
                            f"[dim]Ошибка: [red]{processed_filename}[/red][/dim]")
                        conversion_errors += 1
                        processing_stats.conversion_errors += 1
                        logger.error(
                            f"Ошибка при конвертации {file_path} в KML: {error_message}")

                except Exception as e:
                    console.print(
                        f"[dim]Критическая ошибка: [red]{filename}[/red][/dim]")
                    conversion_errors += 1
                    processing_stats.conversion_errors += 1
                    logger.error(
                        f"Критическая ошибка при обработке {file_path}: {e}", exc_info=True)
                finally:
                    # Advance progress bar regardless of success/failure for this file
                    progress.advance(task)

    return conversion_errors


def _prepare_worker_args(separated_files: List[Path], config: Config) -> List[Dict[str, Any]]:
    """
    Подготовка аргументов для рабочих процессов.
    
    Args:
        separated_files: Список путей к разделенным Excel файлам
        
    Returns:
        Список словарей с именованными аргументами для process_file_worker
    """
    worker_args = []
    for xlsx_file_path in separated_files:
        # Determine the relative path from the separated base dir
        relative_path = xlsx_file_path.relative_to(
            Path(config.xlsx_output_dir))
        # Construct the corresponding KML output path
        kml_file_rel_path = relative_path.with_suffix('.kml')
        kml_file_abs_path = Path(config.kml_output_dir) / kml_file_rel_path

        worker_args.append({
            'xlsx_file_path': str(xlsx_file_path),
            'kml_file_path': str(kml_file_abs_path),
            'xlsx_output_dir': config.xlsx_output_dir,
            'kml_output_dir': config.kml_output_dir
        })
    return worker_args


def _determine_max_workers(separated_files: List[Path], config: Config) -> int:
    """Определение максимального количества рабочих потоков."""
    if config.max_parallel_workers is not None:
        return min(len(separated_files), config.max_parallel_workers)
    else:
        return min(len(separated_files), multiprocessing.cpu_count())


def _report_conversion_results(separated_files: List[Path], conversion_errors: int, config: Config):
    """Отчет о результатах преобразования."""
    if conversion_errors == 0:
        console.print(Panel(
            f"[bold green]✅ Этап 2 завершен успешно![/bold green]\n\n"
            f"Все {len(separated_files)} файлов успешно преобразованы в KML.\n"
            f"KML файлы: [blue]{config.kml_output_dir}[/blue]",
            title="🎉 Преобразование завершено",
            border_style="green"
        ))
    else:
        successful_files = len(separated_files) - conversion_errors

        # Get log file path for error reference
        log_file_path = "неизвестен"
        if logger.handlers:
            for handler in logger.handlers:
                if hasattr(handler, 'baseFilename'):
                    log_file_path = str(
                        getattr(handler, 'baseFilename', 'неизвестен'))
                    break

        console.print(Panel(
            f"[bold yellow]⚠️ Этап 2 завершен с ошибками[/bold yellow]\n\n"
            f"Успешно преобразовано: [green]{successful_files}[/green] файлов\n"
            f"Ошибок: [red]{conversion_errors}[/red]\n\n"
            f"KML файлы: [blue]{config.kml_output_dir}[/blue]\n"
            f"Лог-файл: [dim]{log_file_path}[/dim]",
            title="⚠️ Преобразование завершено с ошибками",
            border_style="yellow"
        ))


def process_mode_2_single_file(config: Config):
    """Обработка режима 2: Преобразование одного файла."""
    console.print(Panel(
        "[bold cyan]Режим: Преобразование одного файла .xlsx в .kml[/bold cyan]\n\n"
        "[dim]Быстрое преобразование одного файла Excel в формат KML\n"
        "без разделения по регионам.[/dim]",
        title="🚀 Быстрое преобразование",
        border_style="cyan"
    ))

    file_name = choose_file(config)
    if not file_name:
        return

    input_path = Path(file_name)

    # Ensure the output directory exists
    Path(config.single_kml_output_dir).mkdir(parents=True, exist_ok=True)

    # Create KML filename in the output directory
    output_filename = Path(config.single_kml_output_dir) / \
        f"{input_path.stem}.kml"

    # Display processing info
    info_table = Table(show_header=False, box=None)
    info_table.add_column("Параметр", style="bold", width=20)
    info_table.add_column("Значение", style="green")

    info_table.add_row("Входной файл:", input_path.name)
    info_table.add_row("Выходной файл:", str(output_filename))

    console.print(Panel(
        info_table,
        title="ℹ️ Параметры преобразования",
        border_style="blue"
    ))

    try:
        # Initialize statistics for single file mode
        single_stats = ProcessingStats()
        single_stats.regions_detected = 1  # Single file = 1 "region"

        with console.status("[cyan]Преобразование файла в KML...[/cyan]", spinner="dots"):
            # Ensure reading only data, not formulas
            workbook = load_workbook(filename=str(input_path), data_only=True)

            # Use enhanced conversion function that collects statistics
            conversion_result = create_kml_from_coordinates(
                workbook.active,
                output_file=str(output_filename),
                filename=input_path.name
            )

            # Add result to statistics
            single_stats.add_file_result(conversion_result)

            # Count anomaly files
            if conversion_result.anomaly_file_created:
                single_stats.anomaly_files_generated += 1

        # Success message
        success_msg = f"[bold green]✅ Преобразование завершено успешно![/bold green]\n\n"
        success_msg += f"Входной файл: [cyan]{input_path.name}[/cyan]\n"
        success_msg += f"Выходной файл: [blue]{output_filename}[/blue]"

        if conversion_result.anomaly_file_created:
            success_msg += f"\n\n[yellow]📊 Создан файл с аномалиями[/yellow]"

        console.print(Panel(
            success_msg,
            title="🎉 Готово",
            border_style="green"
        ))

        # Display comprehensive statistics for single file
        display_processing_statistics(single_stats)

    except Exception as e:
        console.print(Panel(
            f"[bold red]Ошибка при обработке файла:[/bold red]\n{e}\n\n"
            "[dim]Проверьте, что файл не поврежден и содержит корректные данные.[/dim]",
            title="❌ Ошибка преобразования",
            border_style="red"
        ))
        logger.exception(f"Ошибка в режиме 2 при обработке файла {file_name}")


def show_main_menu() -> str:
    """Display main menu and get user choice."""
    menu_table = Table(show_header=False, box=None, padding=(0, 2))
    menu_table.add_column("№", style="bold cyan", width=3)
    menu_table.add_column("Описание", style="white")
    menu_table.add_column("Действие", style="dim")

    menu_table.add_row(
        "1", "Разделить файл по регионам и преобразовать в KML", "Полный цикл обработки")
    menu_table.add_row(
        "2", "Преобразовать один файл .xlsx в .kml", "Быстрое преобразование")
    menu_table.add_row(
        "3", "Отладочный парсинг строк с координатами", "Тестирование парсера")
    menu_table.add_row("4", "Выход", "Завершить работу")

    console.print(Panel(
        menu_table,
        title="📋 Главное меню",
        border_style="cyan"
    ))

    return Prompt.ask(
        "Выберите режим",
        choices=["1", "2", "3", "4"],
        show_choices=False
    )


def main():
    """Main application entry point."""
    global logger
    
    # Set up logging for the MAIN process here, just once.
    # This ensures logging is configured before any other operations.
    setup_logging(console_level=logging.DEBUG)
    logger = logging.getLogger(__name__)
    
    # Create configuration instance
    config = Config()
    
    display_welcome()

    while True:
        try:
            user_input = show_main_menu()
        except (KeyboardInterrupt, EOFError):
            logger.info("Program terminated by user.")
            console.print("\n[yellow]Работа программы завершена.[/yellow]")
            break

        if user_input == "1":
            process_mode_1_full_processing(config)

        elif user_input == "2":
            process_mode_2_single_file(config)

        elif user_input == "3":
            debug_coordinate_parser()

        elif user_input == "4":
            console.print(Panel(
                "[yellow]Спасибо за использование Excel to KML Converter![/yellow]\n\n"
                "[dim]Программа завершена.[/dim]",
                title="👋 До свидания",
                border_style="yellow"
            ))
            break


if __name__ == '__main__':
    # Support for Windows multiprocessing
    multiprocessing.freeze_support()
    main()
