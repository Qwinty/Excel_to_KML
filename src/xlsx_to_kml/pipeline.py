import logging
import os
import time
from typing import List, Optional, Tuple

import simplekml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pyproj import Transformer

from src.utils import generate_random_color, sort_coordinates, FilenameLoggerAdapter
from src.config import Config

from .models import ConversionResult, Point, ParseError, WaterUsageType, get_water_usage_type, generate_point_name
from .parsing import parse_coordinates
from .io_excel import get_column_indices
from .io_kml import create_kml_point, create_kml_line, create_kml_polygon

logger = logging.getLogger(__name__)


def save_anomalies_to_excel(anomalies: List[dict], original_basename: str, output_directory: str) -> bool:
    """Saves detected anomalies to a separate Excel file in the specified output directory. Returns True on success, False otherwise."""
    if not anomalies:
        return False

    name, ext = os.path.splitext(original_basename)
    output_filename = f"ANO_{name}.xlsx"
    output_path = os.path.join(output_directory, output_filename)

    logger.info(f"Saving {len(anomalies)} anomalies to '{output_path}'...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    headers = ["Строка в оригинальном файле", "№ п/п", "Причина", "Координаты"]
    ws.append(headers)

    for anomaly in anomalies:
        ws.append([
            anomaly.get("row_index", "N/A"),
            anomaly.get("main_name", "N/A"),
            anomaly.get("reason", "N/A"),
            anomaly.get("coords_str", "N/A"),
        ])

    for col_idx, col in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row,
                     min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value is not None:
                try:
                    value_str = str(cell.value)
                    if len(value_str) > max_length:
                        max_length = len(value_str)
                except Exception as e:
                    logger.warning(
                        f"Could not determine length for cell value {cell.value} in column {column_letter}: {e}")
        adjusted_width = (max_length + 2)
        if adjusted_width > 64:
            adjusted_width = 64
        ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(output_path)
        logger.info(f"Anomalies successfully saved to '{output_path}'.")
        return True
    except Exception as e:
        logger.error(
            f"Failed to save anomalies to '{output_path}': {e}", exc_info=True)
        print(
            f"[bold red]Ошибка при сохранении файла аномалий '{output_path}': {e}[/bold red]")
        return False


def create_kml_from_coordinates(
    sheet,
    output_file: str = "output.kml",
    sort_numbers: Optional[List[int]] = None,
    filename: Optional[str] = None,
    transformers: Optional[dict[str, Transformer]] = None,
    proj4_path: str = "data/proj4.json",
    config: Config | None = None,
) -> ConversionResult:
    if config is None:
        config = Config()
    start_time = time.time()

    file_logger = FilenameLoggerAdapter(logger, filename)

    stats = ConversionResult(
        filename=filename or os.path.basename(output_file))
    kml = simplekml.Kml()
    indices = get_column_indices(sheet, config=config)
    anomalies_list: List[dict] = []

    min_row = config.excel_default_data_start_row
    for row in sheet.iter_rows(min_row=config.excel_header_scan_min_row, max_row=config.excel_header_scan_max_row):
        cell = row[indices["coord"]] if indices["coord"] != -1 else None
        value = cell.value
        if isinstance(value, str) and ('м.' in value or '"' in value):
            min_row = cell.row
            break

    for row_idx, row in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), start=min_row):
        coords_str = row[indices["coord"]] if indices["coord"] != -1 else None
        if not isinstance(coords_str, str) or not coords_str.strip():
            continue

        stats.total_rows += 1

        main_name = row[indices["name"]
                        ] if indices["name"] != -1 else f"Row {row_idx}"
        file_logger.info(f"------------")

        try:
            coords_array: List[Point] = parse_coordinates(
                coords_str, transformers=transformers, proj4_path=config.proj4_path, config=config
            )
        except ParseError as e:
            error_reason = str(e)
            file_logger.warning(
                f"Строка {row_idx} (№ п/п {main_name}) пропущена из-за ошибки парсинга: {error_reason}")

            stats.failed_rows += 1
            stats.error_reasons.append(error_reason)
            anomalies_list.append({
                "row_index": row_idx,
                "main_name": main_name,
                "reason": error_reason,
                "coords_str": coords_str,
            })
            continue

        if not coords_array:
            file_logger.debug(
                f"Строка {row_idx} (№ п/п {main_name}) не содержит валидных координат для KML.")
            stats.successful_rows += 1
            continue

        stats.successful_rows += 1
        file_logger.info(
            f"Строка {row_idx} (№ п/п {main_name}): Распознано {len(coords_array)} точек.")

        if coords_array:
            color = generate_random_color()

            desc_parts: List[str] = []
            for key, column_name in [
                ("organ", "Уполномоченный орган"),
                ("additional_name", "Наименование водного объекта"),
                ("goal", "Цель водопользования"),
                ("vid", "Вид водопользования"),
                ("owner", "Владелец"),
                ("inn", "ИНН"),
                ("start_date", "Дата начала водопользования"),
                ("end_date", "Дата окончания водопользования"),
                ("coord", "Место водопользования"),
            ]:
                if indices[key] != -1 and row[indices[key]]:
                    if key in ["start_date", "end_date"] and hasattr(row[indices[key]], "date"):
                        date_value = row[indices[key]].date()
                        desc_parts.append(f"{column_name}: {date_value}")
                    elif key in ["start_date", "end_date"] and isinstance(row[indices[key]], str):
                        date_str = row[indices[key]].split(" ")[0]
                        desc_parts.append(f"{column_name}: {date_str}")
                    else:
                        desc_parts.append(
                            f"{column_name}: {row[indices[key]]}")

            description = "\n".join(desc_parts)
            description += "\n == Разработано RUDI.ru =="

            # Определяем тип водопользования один раз
            goal_text = row[indices["goal"]] if indices["goal"] != -1 else ""
            water_type = get_water_usage_type(goal_text)

            skip_terms = config.pipeline_skip_terms

            # Проверяем, можно ли создать полигон
            if len(coords_array) > 3 and not any(term in goal_text for term in skip_terms):
                file_logger.debug(
                    f"Строка {row_idx} (№ п/п {main_name}): Создание полигона")

                if (sort_numbers and int(main_name) in sort_numbers) or len(coords_array) == 4:
                    sorted_coords = sort_coordinates(
                        [(p.lon, p.lat) for p in coords_array])
                else:
                    sorted_coords = [(p.lon, p.lat) for p in coords_array]

                create_kml_polygon(
                    kml, name=f"№ п/п {main_name}", coords=sorted_coords, description=description, color=color, config=config)

                # Проверяем, можно ли создать линию (только для прочих типов водопользования)
            elif (len(coords_array) > 2
                  and all(p.name.startswith("точка") for p in coords_array)
                  and water_type == WaterUsageType.OTHER):
                file_logger.debug(
                    f"Строка {row_idx} (№ п/п {main_name}): Создание линии")
                create_kml_line(kml, name=f"№ п/п {main_name}", coords=[
                                (p.lon, p.lat) for p in coords_array], description=description, color=color, config=config)

            # Создаем отдельные точки
            else:
                index = 1
                for p in coords_array:
                    file_logger.debug(
                        f"  Точка: {p.name} ({p.lat}, {p.lon})")

                    full_name = generate_point_name(
                        main_name, water_type, index, p.name)
                    create_kml_point(
                        kml, full_name, (p.lon, p.lat), description, color, config=config)
                    index += 1

    kml.save(output_file)

    if anomalies_list and output_file:
        output_dir = os.path.dirname(output_file) or '.'
        original_basename = os.path.basename(output_file)
        stats.anomaly_file_created = save_anomalies_to_excel(
            anomalies_list, original_basename, output_dir)
        stats.anomaly_rows = len(anomalies_list)
    elif anomalies_list and not output_file:
        file_logger.warning(
            "Anomalies were detected, but the original filename was not provided. Anomalies will not be saved to a separate file.")
        stats.anomaly_rows = len(anomalies_list)

    stats.processing_time = time.time() - start_time
    return stats
