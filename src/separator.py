import openpyxl
import re
from pathlib import Path
from openpyxl.utils import get_column_letter  # type: ignore[attr-defined]
from openpyxl.styles import Font
import logging  # Импортируем модуль логирования
import time


# --- Configuration ---
INPUT_FILE = './input/BigTable (Trimmed).xlsx'
OUTPUT_DIR = 'output/xlsx'
HEADER_ROW_COUNT = 5
FULL_WIDTH_MERGE_COLUMNS = (1, 7)  # Столбцы A-G
# --- End Configuration ---


# --- Setup Logging using the utility function ---
# Get logger for this module (configuration will be handled by main.py)
logger = logging.getLogger(__name__)
# --- End Logging Setup ---


# --- Helper Functions ---

def sanitize_filename(name):
    """Удаляет недопустимые символы из имени файла/папки."""
    name = str(name).strip()
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', ' ', name)
    name = name.strip('_')
    name = re.sub(r'_+', '_', name)
    if not name:
        name = "unnamed"
    return name


def copy_column_widths(source_sheet, target_sheet):
    """Копирует ширину столбцов с исходного листа на целевой."""
    for col_idx, dimension in source_sheet.column_dimensions.items():
        if dimension.width:
            target_sheet.column_dimensions[col_idx].width = dimension.width


def copy_merged_cells(source_sheet, target_sheet, header_row_count):
    """Копирует диапазоны объединенных ячеек В ПРЕДЕЛАХ ЗАГОЛОВКА."""
    if hasattr(source_sheet, 'merged_cells') and source_sheet.merged_cells:
        try:
            merged_ranges = source_sheet.merged_cells.ranges
            for merged_range in merged_ranges:
                if merged_range.min_row <= header_row_count:
                    target_sheet.cell(row=merged_range.min_row,
                                      column=merged_range.min_col)
                    target_sheet.cell(row=merged_range.max_row,
                                      column=merged_range.max_col)
                    target_sheet.merge_cells(str(merged_range))
        except ValueError as e:
            logging.warning("Could not merge cells %s: %s",
                            merged_range, e)  # Логируем предупреждение
        except Exception as e_gen:
            logging.error("Error merging cells %s: %s",
                          merged_range, e_gen)  # Логируем ошибку


def copy_row_values(row):
    """Извлекает значения из ячеек строки."""
    return [cell.value for cell in row]


def get_first_non_empty_value(row_values):
    """Находит первое непустое значение в списке значений ячеек строки."""
    return next((str(cell) for cell in row_values if cell and str(cell).strip()), None)


def get_full_row_text(row_values):
    """Объединяет непустые значения ячеек строки в одну строку."""
    return " ".join(filter(None, map(str, row_values))).strip()

# --- Основная логика обработки ---


def split_excel_file_by_merges(input_path, output_base_dir, header_rows_count, merge_cols):
    """
    Разделяет файл Excel, используя строки, объединенные на всю ширину, как основные разделители.
    Оптимизировано: метаданные читаются в обычном режиме, данные — в стриминговом (values_only).
    """
    total_start_time = time.time()
    logging.info("--- Запуск процесса разделения файла ---")

    # --- 1. Чтение метаданных: слияния, ширины, шапка ---
    meta_wb = openpyxl.load_workbook(
        input_path, data_only=True, read_only=False)
    meta_ws = meta_wb.active

    # Все объединённые диапазоны
    all_merged_ranges = list(meta_ws.merged_cells.ranges)

    # Вычисляем строки «полной ширины» (колонны merge_cols)
    min_col_target, max_col_target = merge_cols
    full_width_merged_rows = {}
    for mr in all_merged_ranges:
        if (mr.min_row == mr.max_row
                and mr.min_col == min_col_target
                and mr.max_col == max_col_target):
            val = meta_ws.cell(row=mr.min_row, column=mr.min_col).value
            full_width_merged_rows[mr.min_row] = str(
                val).strip() if val is not None else ""

    # Читаем заголовок (значения)
    header_rows_data = [
        list(row_values)
        for row_values in meta_ws.iter_rows(
            min_row=1, max_row=header_rows_count,
            values_only=True
        )
    ]

    # --- Вставка инструкции ---
    instruction_display = "Инструкция по использованию KML"
    instruction_url = "https://www.rudi.ru/kml-instruction.php"
    num_cols = len(header_rows_data[0]) if header_rows_data else merge_cols[1]
    instruction_row = [instruction_display] + [None] * (num_cols - 1)
    # Вставляем инструкцию как 3-ю строку (индекс 2)
    header_rows_data.insert(2, instruction_row)  # type: ignore[arg-type]

    # Ширины столбцов
    source_col_widths = {
        openpyxl.utils.column_index_from_string(col_letter): dim.width
        for col_letter, dim in meta_ws.column_dimensions.items()
        if dim.width
    }

    # --- Корректировка слияний в шапке ---
    header_merged_ranges = []
    # 1. Добавляем новое слияние для строки с инструкцией (строка 3)
    min_col_letter = get_column_letter(merge_cols[0])
    max_col_letter = get_column_letter(merge_cols[1])
    header_merged_ranges.append(f"{min_col_letter}3:{max_col_letter}3")

    # 2. Обрабатываем существующие слияния из оригинальной шапки
    for rng in all_merged_ranges:
        # Работаем только со слияниями из оригинальной шапки
        if rng.min_row <= header_rows_count:
            new_min_row, new_max_row = rng.min_row, rng.max_row

            # Сдвигаем вниз все, что было на 3-й строке и ниже
            if new_min_row >= 3:
                new_min_row += 1
                new_max_row += 1

            # Собираем новую координату в виде строки
            new_coord = f"{get_column_letter(rng.min_col)}{new_min_row}:{get_column_letter(rng.max_col)}{new_max_row}"
            header_merged_ranges.append(new_coord)

    meta_wb.close()

    # --- 2. Потоковое чтение данных (стриминг) ---
    data_wb = openpyxl.load_workbook(
        input_path, data_only=True, read_only=True)
    ws = data_wb.active

    output_path = Path(output_base_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    logging.info("Выходная папка: %s", output_path)

    # Контекст
    current_bvu_name = None
    current_region_name = None
    current_region_data = []
    current_bvu_folder_path = None

    processed_rows_count = 0
    files_saved_count = 0
    data_rows_collected = 0

    # Предварительно подготовим множества ключевых слов
    BKU_WORDS = {'бву', 'комитет', 'департамент'}
    REGION_WORDS = {'область', 'край', 'автономная', 'республика',
                    'округ', 'севастополь', 'москва', 'санкт-петербург'}

    logging.info("Обработка строк начиная с %d...", header_rows_count + 1)
    iter_start = time.time()

    for row_idx, row_values in enumerate(
        ws.iter_rows(
            min_row=header_rows_count + 1,
            values_only=True
        ),
        start=header_rows_count + 1
    ):
        processed_rows_count += 1

        # Преобразуем tuple → список (для удобства последующей записи)
        # (если вы не модифицируете row_values, можно оставить tuple)
        row_list = list(row_values)

        if row_idx in full_width_merged_rows:
            # Это «заголовочная» строка полной ширины
            merged_text = full_width_merged_rows[row_idx]
            mt_low = merged_text.lower()

            is_region_end = merged_text.startswith(
                "Итого действующих документов по субъекту РФ:")
            is_bvu_end = merged_text.startswith(
                "Итого действующих документов по зоне деятельности БВУ:")

            if is_region_end or is_bvu_end:
                # Финиш региона (и, возможно, БВУ)
                if current_bvu_name and current_region_name and current_region_data:
                    save_region_file_optimized(
                        header_rows_data, current_region_data,
                        current_bvu_folder_path, current_region_name,
                        source_col_widths, header_merged_ranges
                    )
                    files_saved_count += 1
                current_region_data = []
                current_region_name = None

                if is_bvu_end:
                    current_bvu_name = None
                    current_bvu_folder_path = None

            elif any(w in mt_low for w in BKU_WORDS):
                # Новый БВУ
                if current_bvu_name and current_region_name and current_region_data:
                    save_region_file_optimized(
                        header_rows_data, current_region_data,
                        current_bvu_folder_path, current_region_name,
                        source_col_widths, header_merged_ranges
                    )
                    files_saved_count += 1

                current_bvu_name = sanitize_filename(merged_text)
                current_bvu_folder_path = output_path / current_bvu_name
                current_bvu_folder_path.mkdir(parents=True, exist_ok=True)
                current_region_name = None
                current_region_data = []

            elif any(w in mt_low for w in REGION_WORDS):
                # Новый Регион
                if current_bvu_name and current_region_name and current_region_data:
                    save_region_file_optimized(
                        header_rows_data, current_region_data,
                        current_bvu_folder_path, current_region_name,
                        source_col_widths, header_merged_ranges
                    )
                    files_saved_count += 1

                current_region_name = sanitize_filename(merged_text)
                current_region_data = []

            else:
                logging.warning(
                    "Неопознанная merged-строка %d: '%s'", row_idx, merged_text)

            # Сброс счётчика сбора данных
            data_rows_collected = 0

        else:
            # Обычная строка с данными
            if current_bvu_name and current_region_name and row_values[0] is not None:
                current_region_data.append(row_list)
                data_rows_collected += 1
                if data_rows_collected % 500 == 0:
                    logging.debug(
                        "Собрано %d строк для %s/%s",
                        data_rows_collected, current_bvu_name, current_region_name
                    )

        # Лог прогресса
        if processed_rows_count % 2000 == 0:
            elapsed = time.time() - iter_start
            logging.info(
                "  Обработано %d строк (%.2f строк/сек)",
                processed_rows_count, processed_rows_count / elapsed if elapsed > 0 else 0
            )

    # После цикла: сохраняем остатки
    if current_bvu_name and current_region_name and current_region_data:
        save_region_file_optimized(
            header_rows_data, current_region_data,
            current_bvu_folder_path, current_region_name,
            source_col_widths, header_merged_ranges
        )
        files_saved_count += 1

    data_wb.close()
    logging.info(
        "Завершено. Файлов сохранено: %d. Всего времени: %.2f сек",
        files_saved_count, time.time() - total_start_time
    )


def save_region_file_optimized(header_data, region_data, bvu_folder_path, region_name,
                               source_col_widths, header_merged_ranges):
    """
    Создает и сохраняет новый Excel-файл для указанного региона в режиме write_only (streaming).
    После сохранения повторно открывает файл для применения объединений заголовка.
    """
    if not region_data:
        logging.info(
            "    Пропуск сохранения для '%s' — нет данных.", region_name)
        return

    if not bvu_folder_path or not region_name:
        logging.warning("    Пропуск сохранения — неверный путь '%s' или имя региона '%s'.",
                        bvu_folder_path, region_name)
        return
    if not bvu_folder_path.exists():
        try:
            bvu_folder_path.mkdir(parents=True, exist_ok=True)
            logging.info("    Создана папка БВУ: %s", bvu_folder_path)
        except Exception as e:
            logging.error("    Не удалось создать папку '%s': %s",
                          bvu_folder_path, e)
            return

    filename = f"{region_name}.xlsx"
    filepath = bvu_folder_path / filename

    try:
        start_time = time.time()
        logging.info("      Начало стриминговой записи файла: %s", filepath)

        # 1) Стриминговая запись: write_only
        wb_stream = openpyxl.Workbook(write_only=True)
        ws_stream = wb_stream.create_sheet()

        # 1.1) Шапка без объединений
        for row in header_data:
            ws_stream.append(row)
        # 1.2) Ширины колонок
        for col_idx, width in source_col_widths.items():
            ws_stream.column_dimensions[get_column_letter(
                col_idx)].width = width
        # 1.3) Данные региона
        for row in region_data:
            ws_stream.append(row)

        wb_stream.save(filepath)
        elapsed = time.time() - start_time
        logging.info("      Стриминговая запись завершена (%.2f сек)", elapsed)

        # 2) Применение объединений через повторное открытие
        if header_merged_ranges:
            wb_norm = openpyxl.load_workbook(filepath)
            ws_norm = wb_norm.active
            for rng in header_merged_ranges:
                try:
                    ws_norm.merge_cells(rng)
                except Exception as e:
                    logging.warning(
                        "        Не удалось объединить %s: %s", rng, e)

            # --- Добавление гиперссылки на строку инструкции ---
            try:
                link_cell = ws_norm.cell(row=3, column=1)
                link_cell.value = "Инструкция по использованию KML"
                link_cell.hyperlink = "https://www.rudi.ru/kml-instruction.php"
                link_cell.font = Font(color="0000FF", underline="single")
            except Exception as e:
                logging.warning(
                    "        Не удалось добавить гиперссылку: %s", e)

            wb_norm.save(filepath)

    except Exception as e:
        logging.exception(
            "      Ошибка при сохранении файла %s: %s", filepath, e)


def save_region_file(header_data, region_data, bvu_folder_path, region_name, source_sheet, header_row_count):
    """ Создает и сохраняет новый Excel файл для указанного региона. """
    if not region_data:
        logging.info(
            "    Пропуск сохранения для '%s' - не найдено данных для этого раздела.", region_name)
        return

    if not bvu_folder_path or not region_name:
        logging.warning(
            "    Пропуск сохранения - Некорректный путь БВУ ('%s') или имя Региона ('%s').", bvu_folder_path, region_name)
        return

    if not bvu_folder_path.exists():
        logging.warning(
            "    Папка БВУ '%s' не существует. Попытка создания.", bvu_folder_path)
        try:
            bvu_folder_path.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logging.error(
                "      Ошибка создания папки '%s': %s. Невозможно сохранить файл.", bvu_folder_path, e)
            return

    filename = f"{region_name}.xlsx"
    filepath = bvu_folder_path / filename

    logging.info("      Создание файла: %s", filepath)
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Запись заголовка и данных
    for header_row in header_data:
        new_sheet.append(header_row)
    for data_row in region_data:
        new_sheet.append(data_row)

    # Копирование форматирования
    copy_column_widths(source_sheet, new_sheet)
    copy_merged_cells(source_sheet, new_sheet, header_row_count)

    # Сохранение
    try:
        new_wb.save(filepath)
        logging.info("      Файл успешно сохранен: %s", filepath)
    except Exception as e:
        # Используем exception для автоматического добавления трейсбека в лог
        logging.exception("      Ошибка при сохранении файла %s", filepath)


# --- Запуск скрипта ---
if __name__ == "__main__":
    # Настройка выполнена в начале файла

    # --- КРИТИЧЕСКИ ВАЖНО: Установите правильное значение HEADER_ROW_COUNT ---
    actual_header_rows = 5  # <--- ИЗМЕНИТЕ ПРИ НЕОБХОДИМОСТИ

    # --- Проверьте и при необходимости измените диапазон столбцов для проверки слияния ---
    actual_merge_columns = (1, 7)  # (A, G)

    # Запуск основной функции
    split_excel_file_by_merges(
        INPUT_FILE, OUTPUT_DIR, actual_header_rows, actual_merge_columns)

    # Добавляем сообщение в лог об окончании работы
    logging.shutdown()  # Корректно закрываем лог-файл
