import logging
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn, SpinnerColumn

from src.config import Config
from src.stats import ProcessingStats, display_processing_statistics
from src.ui import console, choose_file, choose_demo_files_mode, choose_demo_percentage
from src.separator import split_excel_file_by_merges
from src.workers import initialize_worker_logging, process_file_worker
from src.xlsx_to_kml.models import ConversionResult


logger = logging.getLogger(__name__)


def process_mode_1_full_processing(config: Config) -> None:
    console.print(Panel(
        "[bold cyan]Режим: Разделение файла и преобразование в KML[/bold cyan]\n\n"
        "[dim]Этот режим выполнит полный цикл обработки:\n"
        "1. Разделение файла по регионам\n"
        "2. Преобразование каждого региона в KML[/dim]",
        title="🔄 Полная обработка",
        border_style="cyan"
    ))

    input_file = choose_file(config)
    if not input_file:
        return

    input_filename = Path(input_file).name

    processing_stats = ProcessingStats()

    # Stage 1: Separation
    separation_success = _process_file_separation(
        input_file, input_filename, processing_stats, config)

    # Stage 2: KML Conversion
    if separation_success:
        _process_kml_conversion(processing_stats, config)
        display_processing_statistics(processing_stats)
        _log_processing_summary(processing_stats)


def _process_file_separation(input_file: str, input_filename: str, processing_stats: ProcessingStats, config: Config) -> bool:
    separation_success = False

    console.print("[cyan]🔄 Этап 1: Разделение файла по регионам...[/cyan]")

    try:
        Path(config.xlsx_output_dir).mkdir(parents=True, exist_ok=True)
        logger.info(
            f"Создана папка для разделенных XLSX: {config.xlsx_output_dir}")

        split_excel_file_by_merges(
            input_path=input_file,
            output_base_dir=config.xlsx_output_dir,
            header_rows_count=config.header_rows_count,
            merge_cols=config.merge_columns
        )

        separated_files = list(Path(config.xlsx_output_dir).rglob('*.xlsx'))
        processing_stats.regions_detected = len(separated_files)
        processing_stats.files_created = [str(f) for f in separated_files]

        separation_success = True

    except Exception as e:
        console.print(Panel(
            f"[bold red]Ошибка на этапе разделения:[/bold red]\n{e}\n\n"
            "[dim]Проверьте, что файл не открыт в Excel и доступен для чтения.[/dim]",
            title="❌ Ошибка этапа 1",
            border_style="red"
        ))
        logger.exception(
            f"Ошибка в режиме 1 (Разделение) при обработке файла {input_file}")

    if separation_success:
        console.print(Panel(
            f"[bold green]✅ Этап 1 завершен успешно[/bold green]\n\n"
            f"Файл '[cyan]{input_filename}[/cyan]' успешно разделен.\n"
            f"Разделенные XLSX файлы: [blue]{config.xlsx_output_dir}[/blue]",
            title="🎉 Разделение завершено",
            border_style="green"
        ))

    return separation_success


def _process_kml_conversion(processing_stats: ProcessingStats, config: Config) -> None:
    console.print(Panel(
        "[bold cyan]Этап 2: Преобразование разделенных файлов в KML[/bold cyan]\n\n"
        "[dim]Поиск разделенных файлов и преобразование в формат KML...[/dim]",
        title="🔄 Этап 2",
        border_style="cyan"
    ))

    separated_files = list(Path(config.xlsx_output_dir).rglob('*.xlsx'))

    if not separated_files:
        console.print(Panel(
            f"[yellow]Не найдено файлов *.xlsx для преобразования в KML в директории '{config.xlsx_output_dir}' и ее подпапках.[/yellow]",
            title="⚠️ Предупреждение",
            border_style="yellow"
        ))
        return

    console.print(
        f"[green]✓ Найдено {len(separated_files)} файлов .xlsx для преобразования.[/green]")

    Path(config.kml_output_dir).mkdir(parents=True, exist_ok=True)
    logger.info(f"Создана базовая папка для KML: {config.kml_output_dir}")

    conversion_errors = _run_parallel_conversion(
        separated_files, processing_stats, config)
    _report_conversion_results(separated_files, conversion_errors, config)


def _run_parallel_conversion(separated_files: List[Path], processing_stats: ProcessingStats, config: Config) -> int:
    conversion_errors = 0

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total} файлов)"),
        TimeRemainingColumn(),
        console=console,
        transient=False
    ) as progress:
        task = progress.add_task(
            "Преобразование в KML...", total=len(separated_files))

        worker_args = _prepare_worker_args(separated_files, config)
        max_workers = _determine_max_workers(separated_files, config)

        console.print(
            f"[dim]Запуск параллельной обработки с {max_workers} потоками...[/dim]")
        console.print(
            f"[dim]DEBUG/WARNING сообщения подавлены в консоли для повышения производительности[/dim]")

        # ensure correct import in subprocess on Windows
        from src.workers import process_file_worker

        with ProcessPoolExecutor(
            max_workers=max_workers,
            initializer=initialize_worker_logging
        ) as executor:
            future_to_file = {
                executor.submit(process_file_worker, **args): args['xlsx_file_path']
                for args in worker_args
            }

            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                filename = Path(file_path).name

                try:
                    success, processed_filename, conversion_result, error_message = future.result()

                    if success:
                        console.print(
                            f"[dim]Завершено: [green]{processed_filename}[/green][/dim]")
                        if conversion_result is not None:
                            processing_stats.add_file_result(conversion_result)
                            if conversion_result.anomaly_file_created:
                                processing_stats.anomaly_files_generated += 1
                    else:
                        console.print(
                            f"[dim]Ошибка: [red]{processed_filename}[/red][/dim]")
                        conversion_errors += 1
                        processing_stats.conversion_errors += 1
                        logger.error(
                            f"Ошибка при конвертации {file_path} в KML: {error_message}")

                except Exception as e:
                    console.print(
                        f"[dim]Критическая ошибка: [red]{filename}[/red][/dim]")
                    conversion_errors += 1
                    processing_stats.conversion_errors += 1
                    logger.error(
                        f"Критическая ошибка при обработке {file_path}: {e}", exc_info=True)
                finally:
                    progress.advance(task)

    return conversion_errors


def _prepare_worker_args(separated_files: List[Path], config: Config) -> List[Dict[str, Any]]:
    worker_args: List[Dict[str, Any]] = []
    for xlsx_file_path in separated_files:
        relative_path = xlsx_file_path.relative_to(
            Path(config.xlsx_output_dir))
        kml_file_rel_path = relative_path.with_suffix('.kml')
        kml_file_abs_path = Path(config.kml_output_dir) / kml_file_rel_path

        worker_args.append({
            'xlsx_file_path': str(xlsx_file_path),
            'kml_file_path': str(kml_file_abs_path),
            'xlsx_output_dir': config.xlsx_output_dir,
            'kml_output_dir': config.kml_output_dir
        })
    return worker_args


def _determine_max_workers(separated_files: List[Path], config: Config) -> int:
    if config.max_parallel_workers is not None:
        return min(len(separated_files), config.max_parallel_workers)
    else:
        return min(len(separated_files), multiprocessing.cpu_count())


def _report_conversion_results(separated_files: List[Path], conversion_errors: int, config: Config) -> None:
    if conversion_errors == 0:
        console.print(Panel(
            f"[bold green]✅ Этап 2 завершен успешно![/bold green]\n\n"
            f"Все {len(separated_files)} файлов успешно преобразованы в KML.\n"
            f"KML файлы: [blue]{config.kml_output_dir}[/blue]",
            title="🎉 Преобразование завершено",
            border_style="green"
        ))
    else:
        successful_files = len(separated_files) - conversion_errors

        log_file_path = "неизвестен"
        logger_root = logging.getLogger()
        if logger_root.handlers:
            for handler in logger_root.handlers:
                if hasattr(handler, 'baseFilename'):
                    log_file_path = str(
                        getattr(handler, 'baseFilename', 'неизвестен'))
                    break

        console.print(Panel(
            f"[bold yellow]⚠️ Этап 2 завершен с ошибками[/bold yellow]\n\n"
            f"Успешно преобразовано: [green]{successful_files}[/green] файлов\n"
            f"Ошибок: [red]{conversion_errors}[/red]\n\n"
            f"KML файлы: [blue]{config.kml_output_dir}[/blue]\n"
            f"Лог-файл: [dim]{log_file_path}[/dim]",
            title="⚠️ Преобразование завершено с ошибками",
            border_style="yellow"
        ))


def process_mode_2_single_file(config: Config) -> None:
    console.print(Panel(
        "[bold cyan]Режим: Преобразование одного файла .xlsx в .kml[/bold cyan]\n\n"
        "[dim]Быстрое преобразование одного файла Excel в формат KML\n"
        "без разделения по регионам.[/dim]",
        title="🚀 Быстрое преобразование",
        border_style="cyan"
    ))

    file_name = choose_file(config)
    if not file_name:
        return

    input_path = Path(file_name)
    Path(config.single_kml_output_dir).mkdir(parents=True, exist_ok=True)
    output_filename = Path(config.single_kml_output_dir) / \
        f"{input_path.stem}.kml"

    from rich.table import Table

    info_table = Table(show_header=False, box=None)
    info_table.add_column("Параметр", style="bold", width=20)
    info_table.add_column("Значение", style="green")
    info_table.add_row("Входной файл:", input_path.name)
    info_table.add_row("Выходной файл:", str(output_filename))

    console.print(
        Panel(info_table, title="ℹ️ Параметры преобразования", border_style="blue"))

    from openpyxl import load_workbook
    from src.xlsx_to_kml import create_kml_from_coordinates, get_transformers
    from src.stats import ProcessingStats, display_processing_statistics

    try:
        single_stats = ProcessingStats()
        single_stats.regions_detected = 1

        with console.status("[cyan]Преобразование файла в KML...[/cyan]", spinner="dots"):
            workbook = load_workbook(filename=str(input_path), data_only=True)
            # Load transformers lazily (cached in current process)
            transformers = None
            try:
                transformers = get_transformers()
            except Exception:
                transformers = None
            conversion_result = create_kml_from_coordinates(
                workbook.active,
                output_file=str(output_filename),
                filename=input_path.name,
                transformers=transformers,
                config=config
            )

            single_stats.add_file_result(conversion_result)
            if conversion_result.anomaly_file_created:
                single_stats.anomaly_files_generated += 1

        # Build final status message considering possible warnings/errors during saving anomalies
        had_anomalies = conversion_result.anomaly_rows > 0
        anomaly_save_failed = had_anomalies and not conversion_result.anomaly_file_created
        had_parsing_failures = conversion_result.failed_rows > 0

        if anomaly_save_failed or had_parsing_failures:
            status_header = "[bold yellow]⚠️ Преобразование завершено с предупреждениями[/bold yellow]\n\n"
            panel_title = "⚠️ Преобразование завершено с предупреждениями"
            panel_style = "yellow"
        else:
            status_header = "[bold green]✅ Преобразование завершено успешно![/bold green]\n\n"
            panel_title = "🎉 Готово"
            panel_style = "green"

        success_msg = status_header
        success_msg += f"Входной файл: [cyan]{input_path.name}[/cyan]\n"
        success_msg += f"Выходной файл: [blue]{output_filename}[/blue]"

        if had_anomalies and conversion_result.anomaly_file_created:
            success_msg += "\n\n[yellow]📊 Создан файл с аномалиями[/yellow]"
        elif anomaly_save_failed:
            success_msg += "\n\n[bold red]❌ Не удалось сохранить файл аномалий. Возможно, файл уже открыт или недостаточно прав на запись.[/bold red]"

        console.print(Panel(success_msg, title=panel_title,
                      border_style=panel_style))

        display_processing_statistics(single_stats)
        _log_processing_summary(single_stats)

    except Exception as e:
        console.print(Panel(
            f"[bold red]Ошибка при обработке файла:[/bold red]\n{e}\n\n"
            "[dim]Проверьте, что файл не поврежден и содержит корректные данные.[/dim]",
            title="❌ Ошибка преобразования",
            border_style="red"
        ))
        logger.exception(f"Ошибка в режиме 2 при обработке файла {file_name}")


def _log_processing_summary(stats: ProcessingStats) -> None:
    """Log a plain-text summary of processing statistics to file logs.

    Mirrors the key numbers shown in the Rich summary panel so they are
    preserved in the log files.
    """
    try:
        totals = stats.get_total_stats()
        total_rows = totals.get('total_rows', 0)
        successful_rows = totals.get('successful_rows', 0)
        success_rate = (successful_rows / total_rows *
                        100) if total_rows > 0 else 0.0

        # Format processing time similar to stats display
        processing_time = stats.get_processing_time()
        if processing_time < 60:
            time_str = f"{processing_time:.1f}с"
        else:
            minutes = int(processing_time // 60)
            seconds = int(processing_time % 60)
            time_str = f"{minutes}м {seconds}с"

        lines: List[str] = []
        lines.append(f"Файлов обнаружено: {stats.regions_detected} регионов")
        if stats.anomaly_files_generated > 0:
            lines.append(
                f"Файлы с аномалиями: {stats.anomaly_files_generated} файла")
        lines.append(
            f"Объектов обработано: {total_rows} строк -> {successful_rows} успешно ({success_rate:.1f}%)"
        )
        lines.append(f"Время обработки: {time_str}")

        logger.info("\n".join(["Сводка обработки:"] + lines))
    except Exception:
        # Do not let logging issues affect the main flow
        logger.debug(
            "Не удалось записать сводку обработки в лог.", exc_info=True)


def process_mode_3_demo_maps(config: Config) -> None:
    """Process demo maps mode - create demo KML files with a percentage of objects."""
    console.print(Panel(
        "[bold cyan]Режим: Создание демо-карт[/bold cyan]\n\n"
        "[dim]Создание демо-версий KML карт с ограниченным количеством объектов\n"
        "из разделенных файлов xlsx.[/dim]",
        title="🎨 Создание демо-карт",
        border_style="cyan"
    ))

    # Get demo percentage
    demo_percentage = choose_demo_percentage()

    # Get files mode (single file or all files)
    files_selection = choose_demo_files_mode(config)
    if not files_selection:
        return

    processing_stats = ProcessingStats()

    if files_selection == "all":
        _process_all_demo_files(demo_percentage, processing_stats, config)
    else:
        _process_single_demo_file(
            files_selection, demo_percentage, processing_stats, config)

    display_processing_statistics(processing_stats)
    _log_processing_summary(processing_stats)


def _process_all_demo_files(demo_percentage: float, processing_stats: ProcessingStats, config: Config) -> None:
    """Process all xlsx files in the output directory for demo conversion."""
    xlsx_dir = Path(config.xlsx_output_dir)
    xlsx_files = list(xlsx_dir.rglob('*.xlsx'))
    # Filter out temp files
    xlsx_files = [f for f in xlsx_files if not f.name.startswith('~$')]

    if not xlsx_files:
        console.print(Panel(
            f"[red]Файлы для обработки не найдены в '{xlsx_dir}'[/red]",
            title="❌ Ошибка",
            border_style="red"
        ))
        return

    console.print(
        f"[green]Найдено {len(xlsx_files)} файлов для создания демо-карт[/green]")

    # Create demo output directory
    Path(config.demo_kml_output_dir).mkdir(parents=True, exist_ok=True)
    logger.info(f"Создана папка для демо KML: {config.demo_kml_output_dir}")

    processing_stats.regions_detected = len(xlsx_files)
    conversion_errors = 0

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total} файлов)"),
        TimeRemainingColumn(),
        console=console,
        transient=False
    ) as progress:
        task = progress.add_task(
            f"Создание демо-карт ({demo_percentage}%)...", total=len(xlsx_files))

        for xlsx_file_path in xlsx_files:
            try:
                # Create output path preserving directory structure
                relative_path = xlsx_file_path.relative_to(xlsx_dir)
                demo_kml_rel_path = relative_path.with_suffix('.kml')
                demo_kml_abs_path = Path(
                    config.demo_kml_output_dir) / demo_kml_rel_path

                # Create parent directories if needed
                demo_kml_abs_path.parent.mkdir(parents=True, exist_ok=True)

                success, conversion_result = _convert_single_file_to_demo_kml(
                    str(xlsx_file_path), str(
                        demo_kml_abs_path), demo_percentage, config
                )

                if success and conversion_result:
                    processing_stats.add_file_result(conversion_result)
                    if conversion_result.anomaly_file_created:
                        processing_stats.anomaly_files_generated += 1
                    console.print(
                        f"[dim]Готово: [green]{xlsx_file_path.name}[/green][/dim]")
                else:
                    conversion_errors += 1
                    processing_stats.conversion_errors += 1
                    console.print(
                        f"[dim]Ошибка: [red]{xlsx_file_path.name}[/red][/dim]")

            except Exception as e:
                conversion_errors += 1
                processing_stats.conversion_errors += 1
                logger.error(
                    f"Ошибка при создании демо-карты для {xlsx_file_path}: {e}", exc_info=True)
                console.print(
                    f"[dim]Критическая ошибка: [red]{xlsx_file_path.name}[/red][/dim]")
            finally:
                progress.advance(task)

    _report_demo_conversion_results(
        len(xlsx_files), conversion_errors, demo_percentage, config)


def _process_single_demo_file(file_path: str, demo_percentage: float, processing_stats: ProcessingStats, config: Config) -> None:
    """Process a single xlsx file for demo conversion."""
    xlsx_file_path = Path(file_path)

    # Create demo output directory
    Path(config.demo_kml_output_dir).mkdir(parents=True, exist_ok=True)

    # Create output path preserving directory structure
    xlsx_dir = Path(config.xlsx_output_dir)
    relative_path = xlsx_file_path.relative_to(xlsx_dir)
    demo_kml_rel_path = relative_path.with_suffix('.kml')
    demo_kml_abs_path = Path(config.demo_kml_output_dir) / demo_kml_rel_path

    # Create parent directories if needed
    demo_kml_abs_path.parent.mkdir(parents=True, exist_ok=True)

    processing_stats.regions_detected = 1

    try:
        with console.status(f"[cyan]Создание демо-карты ({demo_percentage}%)...[/cyan]", spinner="dots"):
            success, conversion_result = _convert_single_file_to_demo_kml(
                str(xlsx_file_path), str(
                    demo_kml_abs_path), demo_percentage, config
            )

        if success and conversion_result:
            processing_stats.add_file_result(conversion_result)
            if conversion_result.anomaly_file_created:
                processing_stats.anomaly_files_generated += 1

            console.print(Panel(
                f"[bold green]✅ Демо-карта создана успешно![/bold green]\n\n"
                f"Исходный файл: [cyan]{xlsx_file_path.name}[/cyan]\n"
                f"Демо-файл: [blue]{demo_kml_abs_path}[/blue]\n"
                f"Процент объектов: [yellow]{demo_percentage}%[/yellow]",
                title="🎉 Готово",
                border_style="green"
            ))
        else:
            processing_stats.conversion_errors = 1
            console.print(Panel(
                f"[bold red]Ошибка при создании демо-карты[/bold red]\n\n"
                f"Файл: [cyan]{xlsx_file_path.name}[/cyan]\n"
                "[dim]Возможно, файл пуст или поврежден[/dim]",
                title="❌ Ошибка",
                border_style="red"
            ))

    except Exception as e:
        processing_stats.conversion_errors = 1
        logger.error(
            f"Ошибка при создании демо-карты для {file_path}: {e}", exc_info=True)
        console.print(Panel(
            f"[bold red]Критическая ошибка при обработке файла[/bold red]\n\n"
            f"Файл: [cyan]{xlsx_file_path.name}[/cyan]\n"
            f"Ошибка: {e}",
            title="❌ Критическая ошибка",
            border_style="red"
        ))


def _convert_single_file_to_demo_kml(xlsx_path: str, kml_path: str, demo_percentage: float, config: Config) -> Tuple[bool, Optional[ConversionResult]]:
    """Convert a single xlsx file to demo KML with specified percentage of objects."""
    try:
        from openpyxl import load_workbook
        from src.xlsx_to_kml import create_kml_from_coordinates, get_transformers

        workbook = load_workbook(filename=xlsx_path, data_only=True)

        # Load transformers
        transformers = None
        try:
            transformers = get_transformers()
        except Exception:
            transformers = None

        conversion_result = create_kml_from_coordinates(
            workbook.active,
            output_file=kml_path,
            filename=Path(xlsx_path).name,
            transformers=transformers,
            config=config,
            demo_percentage=demo_percentage
        )

        # Check if demo file is empty
        if conversion_result.successful_rows == 0:
            logger.warning(
                f"Demo file would be empty for {xlsx_path}, skipping")
            # Remove empty file if it was created
            if Path(kml_path).exists():
                Path(kml_path).unlink()
            return False, None

        return True, conversion_result

    except Exception as e:
        logger.error(
            f"Error converting {xlsx_path} to demo KML: {e}", exc_info=True)
        return False, None


def _report_demo_conversion_results(total_files: int, conversion_errors: int, demo_percentage: float, config: Config) -> None:
    """Report the results of demo conversion."""
    if conversion_errors == 0:
        console.print(Panel(
            f"[bold green]✅ Демо-карты созданы успешно![/bold green]\n\n"
            f"Обработано файлов: {total_files}\n"
            f"Процент объектов: [yellow]{demo_percentage}%[/yellow]\n"
            f"Демо-карты: [blue]{config.demo_kml_output_dir}[/blue]",
            title="🎉 Создание демо-карт завершено",
            border_style="green"
        ))
    else:
        successful_files = total_files - conversion_errors
        console.print(Panel(
            f"[bold yellow]⚠️ Демо-карты созданы с ошибками[/bold yellow]\n\n"
            f"Успешно обработано: [green]{successful_files}[/green] файлов\n"
            f"Ошибок: [red]{conversion_errors}[/red]\n"
            f"Процент объектов: [yellow]{demo_percentage}%[/yellow]\n"
            f"Демо-карты: [blue]{config.demo_kml_output_dir}[/blue]",
            title="⚠️ Создание демо-карт завершено с ошибками",
            border_style="yellow"
        ))
