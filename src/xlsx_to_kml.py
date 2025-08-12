import glob
import json
import logging
import re
import math
import os
import time
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict, cast
from functools import lru_cache
import simplekml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pyproj import CRS, Transformer

# Import necessary functions from utils
from src.utils import generate_random_color, sort_coordinates, FilenameLoggerAdapter

# Get logger for this module (configuration will be handled by main.py)
logger = logging.getLogger(__name__)

# --- Compiled Regex Patterns ---
# Compile regex patterns for better performance
MSK_COORD_PATTERN = re.compile(r'(\d+):\s*([-\d.]+)\s*м\.,\s*([-\d.]+)\s*м\.')
DMS_COORD_PATTERN = re.compile(
    r'(\d+)[°º]\s*(\d+)[\'′΄]\s*(\d+(?:[.,]\d+)?)[\"″′′˝]')
DMS_POINT_PATTERN = re.compile(r'(\d+)[:.]\s*(?=\d+[°º])')
ALT_POINT_PATTERN = re.compile(r'точка\s*(\d+)', re.IGNORECASE)

# --- Statistics Data Structure ---


@dataclass
class ConversionResult:
    """Result of converting a single file to KML."""
    filename: str
    total_rows: int = 0
    successful_rows: int = 0
    failed_rows: int = 0
    anomaly_rows: int = 0
    error_reasons: List[str] = field(default_factory=list)
    processing_time: float = 0.0
    anomaly_file_created: bool = False

    @property
    def success_rate(self) -> float:
        """Calculate success rate as percentage."""
        if self.total_rows == 0:
            return 0.0
        return (self.successful_rows / self.total_rows) * 100

    @property
    def failure_rate(self) -> float:
        """Calculate failure rate as percentage."""
        if self.total_rows == 0:
            return 0.0
        # failed_rows already includes all the problematic rows, no need to add anomaly_rows
        return (self.failed_rows / self.total_rows) * 100


@dataclass(frozen=True)
class Point:
    """Typed representation of a geographic point."""
    name: str
    lon: float
    lat: float


class ParseError(Exception):
    """Raised when coordinate parsing fails with a user-facing reason."""
    pass


def create_transformer(proj4_str: str) -> Transformer:
    """Создает трансформер из заданной строки Proj4 в WGS84."""
    crs = CRS.from_proj4(proj4_str)
    return Transformer.from_crs(crs, "EPSG:4326", always_xy=True)


@lru_cache(maxsize=None)
def get_transformers(proj4_path: str = "data/proj4.json") -> dict[str, Transformer]:
    """Лениво загружает и кэширует словарь трансформеров из файла proj4.json.

    Аргументы:
        proj4_path: Путь к JSON-файлу с описаниями проекций (Proj4).

    Возвращает:
        dict[str, Transformer]: Словарь {имя_системы: Transformer}
    """
    try:
        with open(proj4_path, "r", encoding="utf-8") as f:
            proj4_strings: dict[str, str] = json.load(f)

        # Автосоздание алиасов для МСК, где есть ровно одна зона: "МСК-06 зона 1" -> "МСК-06"
        try:
            zone_key_regex = re.compile(r'^(МСК-[^з]+?)\s+зона\s+\d+\b')
            msk_groups: dict[str, list[str]] = {}

            # Сгруппируем ключи по префиксу до слова "зона"
            for name in list(proj4_strings.keys()):
                match = zone_key_regex.match(name)
                if match:
                    prefix = match.group(1).strip()
                    msk_groups.setdefault(prefix, []).append(name)

            # Для групп, где только одна зона, добавим алиас без слова "зона"
            for prefix, full_names in msk_groups.items():
                if len(full_names) == 1:
                    alias_key = prefix
                    full_key = full_names[0]
                    if alias_key not in proj4_strings:
                        proj4_strings[alias_key] = proj4_strings[full_key]
                        logger.debug(
                            f"Добавлен алиас проекции: '{alias_key}' -> '{full_key}'")
        except Exception as e:
            # Не мешаем запуску, если что-то пойдет не так с алиасами
            logger.warning(f"Не удалось создать алиасы МСК без 'зона': {e}")

        # Создаем трансформеры
        transformers: dict[str, Transformer] = {
            name: create_transformer(proj4)
            for name, proj4 in proj4_strings.items()
        }
        return transformers

    except FileNotFoundError:
        logger.critical(
            f"Critical Error: Could not find '{proj4_path}'. This file is required for coordinate transformations. Ensure it exists.")
        print(
            f"[bold red]Критическая ошибка: Не найден файл '{proj4_path}'.[/bold red]")
        print("[bold red]Этот файл необходим для преобразования координат. Убедитесь, что он находится в папке 'data' рядом с программой.[/bold red]")
        # Не выбрасываем SystemExit здесь, чтобы позволить тестам DMS и другим сценариям работать
        raise
    except json.JSONDecodeError:
        logger.critical(
            f"Critical Error: Could not parse '{proj4_path}'. Check the file format.")
        print(
            f"[bold red]Критическая ошибка: Не удалось прочитать файл '{proj4_path}'. Проверьте формат файла.[/bold red]")
        raise
    except Exception as e:
        logger.critical(
            f"Critical Error: An unexpected error occurred while loading projection data: {e}", exc_info=True)
        print(
            f"[bold red]Критическая ошибка: Непредвиденная ошибка при загрузке данных проекций: {e}[/bold red]")
        raise


def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate the great circle distance between two points on earth (specified in decimal degrees)"""
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * \
        math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371  # Radius of earth in kilometers
    return c * r


def detect_coordinate_anomalies(coordinates: List[Point], threshold_km: float = 20):
    """
    Detect anomalous coordinates in a sequence by looking for points that are
    significantly further away from the majority of other points.

    Args:
        coordinates: List of tuples (name, lon, lat)
        threshold_km: Distance threshold in kilometers

    Returns:
        Tuple of (is_anomalous, reason, anomalous_points)
    """
    if len(coordinates) < 3:
        # Not enough points to detect anomalies
        return False, None, []

    distances = []
    anomalous_points = []

    # Calculate distance from each point to every other point
    for i, point_i in enumerate(coordinates):
        lon_i = point_i.lon
        lat_i = point_i.lat
        point_distances = []
        for j, point_j in enumerate(coordinates):
            if i != j:
                dist = haversine_distance(
                    lat_i, lon_i, point_j.lat, point_j.lon)
                point_distances.append(dist)

        # Calculate the average distance to other points
        avg_distance = sum(point_distances) / len(point_distances)
        distances.append((i, avg_distance))

    # Find points that are much further away from others
    for idx, avg_dist in distances:
        # If a point's average distance to others is larger than the threshold
        if avg_dist > threshold_km:
            p = coordinates[idx]
            anomalous_points.append((idx, p.name, p.lon, p.lat))

    if anomalous_points:
        anomaly_details = ', '.join([f"{point_name} ({lat}, {lon})"
                                     for _, point_name, lon, lat in anomalous_points])
        reason = f"Обнаружены аномальные координаты, значительно удаленные от других"
        return True, reason, anomalous_points

    return False, None, []


def _should_prioritize_dms(coord_str: str) -> bool:
    return 'гск' in coord_str.lower()


def _is_candidate_msk(coord_str: str) -> bool:
    return ((' м.' in coord_str or ', м.' in coord_str or coord_str.endswith('м.'))
            and '°' not in coord_str)


def _validate_wgs84_range(lat: float, lon: float) -> bool:
    return -90 <= lat <= 90 and -180 <= lon <= 180


def parse_msk_coordinates(input_string: str, transformer: Transformer) -> List[Point]:
    """Парсит MSK-координаты и возвращает список точек без детектора аномалий.

    Возвращает [] если MSK-совпадений нет; в случае ошибки бросает ParseError.
    """
    matches = MSK_COORD_PATTERN.findall(input_string)
    if not matches:
        return []

    results: List[Point] = []
    for i, x_str, y_str in matches:
        try:
            x_val = float(x_str)
            y_val = float(y_str)
            if x_val == 0 and y_val == 0:
                continue
            lon, lat = transformer.transform(y_val, x_val)
            if not _validate_wgs84_range(lat, lon):
                reason = (
                    f"Координаты МСК вне допустимого диапазона WGS84 (lat={lat}, lon={lon}) после трансформации.")
                raise ParseError(reason)
            results.append(
                Point(name=f"точка {i}", lon=round(lon, 6), lat=round(lat, 6)))
        except Exception as e:
            reason = (
                f"Ошибка трансформации МСК координат: {e}. Исходные: x='{x_str}', y='{y_str}'.")
            raise ParseError(reason)

    if not results:
        return []
    return results


def _extract_dms_matches(coord_str: str) -> Tuple[List[Dict[str, object]], Dict[int, List[str]]]:
    """Извлекает все DMS-совпадения и номера точек по частям."""
    parts = [p.strip() for p in coord_str.split(';') if p.strip()]
    all_dms_coords: List[Dict[str, object]] = []
    point_numbers_by_part: Dict[int, List[str]] = {}

    for idx, part in enumerate(parts):
        point_numbers = DMS_POINT_PATTERN.findall(part)
        if point_numbers:
            point_numbers_by_part[idx] = point_numbers
        coords_match = DMS_COORD_PATTERN.findall(part)
        for coord in coords_match:
            all_dms_coords.append(
                {'coord': coord, 'part': part, 'part_index': idx})

    return all_dms_coords, point_numbers_by_part


def _dms_tuple_to_decimal(d: Tuple[str, str, str]) -> float:
    """Преобразует кортеж (deg, min, sec) в десятичные градусы."""
    return sum(float(x.replace(',', '.')) / (60 ** k) for k, x in enumerate(d))


def _derive_point_name(part_idx: int, pair_idx: int, mapping: Dict[int, List[str]], part_text: str) -> str:
    name = f"точка {pair_idx + 1}"
    part_point_numbers = mapping.get(part_idx, [])
    if pair_idx < len(part_point_numbers):
        name = f"точка {part_point_numbers[pair_idx]}"
    else:
        alt_match = ALT_POINT_PATTERN.search(part_text)
        if alt_match:
            name = f"точка {alt_match.group(1)}"
    return name


def parse_dms_coordinates(coord_str: str) -> List[Point]:
    """Парсит DMS-координаты и возвращает список точек без детектора аномалий.

    В случае ошибки бросает ParseError.
    """
    all_dms_coords, point_numbers_by_part = _extract_dms_matches(coord_str)

    if not all_dms_coords:
        return []

    if len(all_dms_coords) % 2 != 0:
        reason = (
            f"Нечетное количество найденных ДМС координат ({len(all_dms_coords)}). Ожидается пара (широта, долгота).")
        raise ParseError(reason)

    result: List[Point] = []
    for j in range(0, len(all_dms_coords), 2):
        try:
            lat_info = all_dms_coords[j]
            lon_info = all_dms_coords[j + 1]

            lat_parts = cast(Tuple[str, str, str], lat_info['coord'])
            lon_parts = cast(Tuple[str, str, str], lon_info['coord'])

            lat = _dms_tuple_to_decimal(lat_parts)
            lon = _dms_tuple_to_decimal(lon_parts)

            combined_text = f"{cast(str, lat_info['part'])} {cast(str, lon_info['part'])}"
            if "ЮШ" in combined_text or "S" in combined_text:
                lat = -lat
            if "ЗД" in combined_text or "W" in combined_text:
                lon = -lon

            if not _validate_wgs84_range(lat, lon):
                reason = f"Координаты ДМС вне допустимого диапазона WGS84 (lat={lat}, lon={lon})."
                raise ParseError(reason)

            part_idx = cast(int, lat_info['part_index'])
            pair_idx = j // 2
            point_name = _derive_point_name(
                part_idx, pair_idx, point_numbers_by_part, cast(str, lat_info['part']))

            if lat != 0 or lon != 0:
                result.append(
                    Point(name=point_name, lon=round(lon, 6), lat=round(lat, 6)))
        except Exception as e:
            reason = f"Внутренняя ошибка при обработке пары ДМС: {e}."
            raise ParseError(reason)

    return result


def process_coordinates(input_string: str, transformer: Transformer) -> List[Point]:
    """MSK parsing with anomaly detection. Returns points or raises ParseError."""
    results = parse_msk_coordinates(input_string, transformer)
    if not results:
        return []
    if len(results) >= 3:
        is_anomalous, a_reason, _ = detect_coordinate_anomalies(results)
        if is_anomalous:
            raise ParseError(a_reason)
    return results


def parse_coordinates(
    coord_str: str,
    transformers: Optional[dict[str, Transformer]] = None,
    proj4_path: str = "data/proj4.json",
) -> List[Point]:
    """Парсит строку с координатами и возвращает список `Point`.

    В случае ошибки бросает `ParseError`. Пустой список означает, что валидных координат не найдено.
    """
    if not coord_str or not isinstance(coord_str, str):
        logger.debug("Пустая или нестроковая строка координат")
        return []

    coord_str = coord_str.strip()
    logger.debug(f"1. Исходная строка после удаления пробелов: '{coord_str}'")

    if not coord_str:
        logger.debug(
            "Строка пуста после удаления пробелов. Возвращаем пустой результат.")
        return []

    logger.debug("2. Определение формата координат (детектор)...")

    if _should_prioritize_dms(coord_str):
        logger.debug("  - Обнаружен маркер 'гск'. Приоритет ДМС.")
    else:
        if _is_candidate_msk(coord_str):
            logger.debug("  - Кандидат на МСК-формат. Попытка парсинга МСК.")
            if transformers is None:
                try:
                    transformers = get_transformers(proj4_path)
                except Exception:
                    reason = "Не удалось загрузить описания проекций для МСК."
                    logger.warning(f"{reason} Строка: '{coord_str[:50]}'")
                    raise ParseError(reason)
            for key, transformer in transformers.items():
                if key in coord_str:
                    logger.debug(f"    - Найдена система координат: '{key}'.")
                    # Парсер MSK (без аномалий)
                    msk_points = parse_msk_coordinates(coord_str, transformer)
                    # Отдельный вызов детектора аномалий над готовыми точками
                    if msk_points and len(msk_points) >= 3:
                        is_anomalous, a_reason, _ = detect_coordinate_anomalies(
                            msk_points)
                        if is_anomalous:
                            logger.warning(
                                f"  - Детектор аномалий сообщил: {a_reason}")
                            raise ParseError(a_reason)
                    return msk_points
            reason = "Обнаружены координаты 'м.', но не найдена известная система координат МСК в строке."
            logger.warning(f"{reason} Строка: '{coord_str[:50]}'")
            raise ParseError(reason)

    logger.debug("3. Проверка на наличие маркера ДМС ('°')...")
    if '°' not in coord_str:
        logger.debug(
            "  - Маркер '°' не найден. Предполагается, что в строке нет координат. Возвращаем пустой результат.")
        return []

    logger.debug("  - Маркер '°' найден. Начинается парсинг ДМС.")
    # Парсер DMS (без аномалий)
    dms_points = parse_dms_coordinates(coord_str)

    # Отдельный вызов детектора аномалий над готовыми точками
    if dms_points and len(dms_points) >= 3:
        is_anomalous, reason, _ = detect_coordinate_anomalies(dms_points)
        if is_anomalous:
            logger.warning(f"  - Детектор аномалий сообщил: {reason}")
            raise ParseError(reason)

    logger.debug(
        f"7. Парсинг успешно завершен. Найдено {len(dms_points)} валидных координат.")
    return dms_points


def find_column_index(sheet, target_names: List[str], exact_match: bool = False) -> int:
    """Находит индекс столбца для любого из заданных имен заголовков в строках 1-8.

    Args:
        sheet: Лист Excel для поиска.
        target_names: Список имен заголовков для поиска.
        exact_match: Если True, требуется точное совпадение заголовка, иначе ищет подстроку (по умолчанию False).

    Returns:
        Индекс столбца или -1, если не найдено.
    """
    target_names_lower = [str(name).lower().strip() for name in target_names]
    # В некоторых файлах заголовки могут быть смещены ниже 5-й строки (многострочные шапки)
    for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True):
        for idx, cell in enumerate(row):
            if cell:
                # Нормализуем строку заголовка: приводим к нижнему регистру и убираем лишние пробелы/переводы строк
                cell_str_lower = str(cell).lower().strip()
                for target_name_lower in target_names_lower:
                    if (exact_match and cell_str_lower == target_name_lower) or \
                       (not exact_match and target_name_lower in cell_str_lower):
                        return idx
    return -1


def get_column_indices(sheet) -> dict:
    """Получает индексы всех необходимых столбцов."""
    columns = {
        "name": ["№ п/п"],
        "coord": ["Место водопользования"],
        "organ": ["Уполномоченный орган"],
        "additional_name": ["Наименование водного объекта"],
        "goal": ["Цель водопользования"],
        "vid": ["Вид водопользования"],
        "owner": ["Наименование"],
        "inn": ["ИНН"],
        "start_date": ["Дата начала водопользования"],
        "end_date": ["Дата окончания водопользования", "Дата прекращения действия"]
    }

    # Dictionary to specify which columns need exact matching
    exact_matches = {
        "owner": True
    }

    indices = {}
    for key, value in columns.items():
        # Use exact_match parameter if specified for this key
        exact = exact_matches.get(key, False)
        indices[key] = find_column_index(sheet, value, exact_match=exact)

    # Use the first name from the list for reporting missing columns
    original_names = {key: value[0] for key, value in columns.items()}
    for key, value in indices.items():
        if value == -1:
            logger.debug(
                f"Столбец '{original_names[key]}' (или его альтернативы) не найден.")

    return indices


def create_kml_point(kml, name: str, coords: Tuple[float, float], description: str, color: str) -> None:
    """Создает точку KML с заданными параметрами."""
    point = kml.newpoint(name=name, coords=[coords])
    point.description = description
    point.style.iconstyle.color = color
    point.style.iconstyle.scale = 1.0
    point.style.labelstyle.scale = 0.8


def create_kml_from_coordinates(
    sheet,
    output_file: str = "output.kml",
    sort_numbers: Optional[List[int]] = None,
    filename: Optional[str] = None,
    transformers: Optional[dict[str, Transformer]] = None,
    proj4_path: str = "data/proj4.json",
) -> ConversionResult:
    """Создает KML-файл из листа с координатами и сохраняет аномалии в отдельный файл. 

    Args:
        sheet: Excel worksheet to process
        output_file: Path for output KML file
        sort_numbers: Optional list of numbers for coordinate sorting
        filename: Name of the source file for statistics

    Returns:
        ConversionResult: Detailed statistics about the conversion process
    """
    start_time = time.time()

    # Create logger adapter with filename for automatic inclusion in log messages
    file_logger = FilenameLoggerAdapter(logger, filename)

    # Initialize statistics
    stats = ConversionResult(
        filename=filename or os.path.basename(output_file)
    )
    kml = simplekml.Kml()
    indices = get_column_indices(sheet)
    anomalies_list = []  # Initialize list to store anomalies

    # Default min_row value
    min_row = 6

    # Check rows 2..5 using iter_rows()
    for row in sheet.iter_rows(min_row=2, max_row=5):
        cell = row[indices["coord"]] if indices["coord"] != -1 else None
        value = cell.value

        if isinstance(value, str) and ('м.' in value or '"' in value):
            min_row = cell.row  # Get the actual row number (3 or 4)
            break  # Stop at first match

    # Используем определенное значение min_row в цикле
    for row_idx, row in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), start=min_row):
        coords_str = row[indices["coord"]] if indices["coord"] != -1 else None
        if not isinstance(coords_str, str) or not coords_str.strip():
            continue

        # Count this as a data row
        stats.total_rows += 1

        main_name = row[indices["name"]
                        ] if indices["name"] != -1 else f"Row {row_idx}"
        file_logger.info(f"------------")

        # Вызываем обновленную функцию парсинга
        try:
            coords_array: List[Point] = parse_coordinates(
                coords_str,
                transformers=transformers,
                proj4_path=proj4_path,
            )
        except ParseError as e:
            error_reason = str(e)
            # Логирование уже произошло внутри parse_coordinates
            file_logger.warning(
                f"Строка {row_idx} (№ п/п {main_name}) пропущена из-за ошибки парсинга: {error_reason}")

            # Update statistics
            stats.failed_rows += 1
            stats.error_reasons.append(error_reason)
            # Note: anomaly_rows will be set later based on whether anomaly file was actually created

            # Add anomaly details to the list
            anomalies_list.append({
                "row_index": row_idx,
                "main_name": main_name,
                "reason": error_reason,
                "coords_str": coords_str
            })
            continue  # Переходим к следующей строке Excel

        # Если coords_array это пустой список, значит парсинг прошел успешно, но валидных координат не найдено.
        # В этом случае просто пропускаем создание KML геометрии для этой строки.
        if not coords_array:
            file_logger.debug(
                f"Строка {row_idx} (№ п/п {main_name}) не содержит валидных координат для KML.")
            # Count as successful parsing even though no coordinates found
            stats.successful_rows += 1
            continue  # Нет точек для добавления в KML

        # Count as successful row
        stats.successful_rows += 1

        file_logger.info(
            f"Строка {row_idx} (№ п/п {main_name}): Распознано {len(coords_array)} точек.")

        if coords_array:  # Убедимся еще раз, что список не пуст
            color = generate_random_color()

            # Подготавливаем описание
            desc = []
            for key, column_name in [
                ("organ", "Уполномоченный орган"),
                ("additional_name", "Наименование водного объекта"),
                ("goal", "Цель водопользования"),
                ("vid", "Вид водопользования"),
                ("owner", "Владелец"),
                ("inn", "ИНН"),
                ("start_date", "Дата начала водопользования"),
                ("end_date", "Дата окончания водопользования"),
                ("coord", "Место водопользования")
            ]:
                if indices[key] != -1 and row[indices[key]]:
                    # Форматируем даты без времени, если это даты начала или окончания водопользования
                    if key in ["start_date", "end_date"] and hasattr(row[indices[key]], "date"):
                        # Если это объект datetime, берем только дату
                        date_value = row[indices[key]].date()
                        desc.append(f"{column_name}: {date_value}")
                    elif key in ["start_date", "end_date"] and isinstance(row[indices[key]], str):
                        # Если это строка, обрезаем время, если оно есть
                        date_str = row[indices[key]].split(" ")[0]
                        desc.append(f"{column_name}: {date_str}")
                    else:
                        desc.append(f"{column_name}: {row[indices[key]]}")

            description = '\n'.join(desc)
            description += "\n == Разработано RUDI.ru =="

            # Проверяем, есть ли 16-й столбец

            # Проверяем, есть ли более 3 точек и 16-й столбец не равен нулю или пуст
            skip_terms = ["Сброс сточных", "Забор (изъятие)"]
            if len(coords_array) > 3 and not any(term in row[indices["goal"]] for term in skip_terms):
                file_logger.debug(
                    f"Строка {row_idx} (№ п/п {main_name}): Создание полигона")
                # Создаем полигон
                polygon = kml.newpolygon(name=f"№ п/п {main_name}")

                # Сортируем координаты только если main_name есть в sort_numbers
                if (sort_numbers and int(main_name) in sort_numbers) or len(coords_array) == 4:
                    sorted_coords = sort_coordinates(
                        [(p.lon, p.lat) for p in coords_array])
                else:
                    sorted_coords = [(p.lon, p.lat) for p in coords_array]

                polygon.outerboundaryis = sorted_coords  # type: ignore
                polygon.style.linestyle.color = color
                polygon.style.linestyle.width = 3
                polygon.style.polystyle.color = simplekml.Color.changealphaint(
                    100, color)
                polygon.description = description
            else:
                # Создаем линию, если есть несколько точек и условия выполнены
                if len(coords_array) > 2 \
                        and all(p.name.startswith("точка") for p in coords_array) \
                        and row[indices["goal"]] != "Сброс сточных вод":
                    file_logger.debug(
                        f"Строка {row_idx} (№ п/п {main_name}): Создание линии")
                    line = kml.newlinestring(name=f"№ п/п {main_name}",
                                             coords=[(p.lon, p.lat) for p in coords_array])
                    line.style.linestyle.color = color
                    line.style.linestyle.width = 3
                    line.description = description
                else:
                    # Создаем отдельные точки, если линия не была создана
                    index = 1
                    for p in coords_array:
                        file_logger.debug(
                            f"  Точка: {p.name} ({p.lat}, {p.lon})")
                        # print(f"{lat}, {lon}")
                        if row[indices["goal"]] == "Сброс сточных вод":
                            full_name = f"№ п/п {main_name} - сброс {index}"
                        else:
                            full_name = f"№ п/п {main_name} - {p.name}" if p.name else f"№ п/п {main_name}"
                        create_kml_point(
                            kml, full_name, (p.lon, p.lat), description, color)
                        index += 1

    kml.save(output_file)

    # Handle anomaly file creation
    if anomalies_list and output_file:
        # Use current dir if output_file has no path
        output_dir = os.path.dirname(output_file) or '.'
        original_basename = os.path.basename(output_file)
        # Capture the return value from save_anomalies_to_excel
        stats.anomaly_file_created = save_anomalies_to_excel(
            anomalies_list, original_basename, output_dir)
        # Set anomaly_rows to the actual number of anomalies found
        stats.anomaly_rows = len(anomalies_list)
    elif anomalies_list and not output_file:
        file_logger.warning(
            "Anomalies were detected, but the original filename was not provided. Anomalies will not be saved to a separate file.")
        stats.anomaly_rows = len(anomalies_list)

    # Finalize statistics
    stats.processing_time = time.time() - start_time

    return stats


def create_kml_from_coordinates_legacy(sheet, output_file: str = "output.kml", sort_numbers: Optional[List[int]] = None) -> bool:
    """Legacy wrapper that maintains backward compatibility. Returns True if anomaly file was created."""
    result = create_kml_from_coordinates(sheet, output_file, sort_numbers)
    return result.anomaly_file_created


def save_anomalies_to_excel(anomalies: List[dict], original_basename: str, output_directory: str) -> bool:
    """Saves detected anomalies to a separate Excel file in the specified output directory. Returns True on success, False otherwise."""
    if not anomalies:
        return False  # Nothing to save

    # Construct the output filename using the original basename
    name, ext = os.path.splitext(original_basename)
    output_filename = f"ANO_{name}.xlsx"
    # Place it in the output directory
    output_path = os.path.join(output_directory, output_filename)

    logger.info(f"Saving {len(anomalies)} anomalies to '{output_path}'...")

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    # Write headers
    headers = ["Строка в оригинальном файле", "№ п/п", "Причина", "Координаты"]
    ws.append(headers)

    # Write anomaly data
    for anomaly in anomalies:
        ws.append([
            anomaly.get("row_index", "N/A"),
            anomaly.get("main_name", "N/A"),
            anomaly.get("reason", "N/A"),
            anomaly.get("coords_str", "N/A"),
        ])

    # Adjust column widths (optional, for better readability)
    for col_idx, col in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row,
                     min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value is not None:
                try:
                    value_str = str(cell.value)
                    if len(value_str) > max_length:
                        max_length = len(value_str)
                except Exception as e:
                    logger.warning(
                        f"Could not determine length for cell value {cell.value} in column {column_letter}: {e}")
        adjusted_width = (max_length + 2)
        if adjusted_width > 64:
            adjusted_width = 64
        ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(output_path)
        logger.info(f"Anomalies successfully saved to '{output_path}'.")
        return True  # Return True on successful save
    except Exception as e:
        logger.error(
            f"Failed to save anomalies to '{output_path}': {e}", exc_info=True)
        print(
            f"[bold red]Ошибка при сохранении файла аномалий '{output_path}': {e}[/bold red]")
        return False  # Return False on error


if __name__ == "__main__":
    # Ensure the logger is at DEBUG level for the console when in debug mode
    # This is a bit of a hack, but it ensures debug messages are seen without changing the global config
    for handler in logger.root.handlers:
        if isinstance(handler, logging.StreamHandler):
            handler.setLevel(logging.DEBUG)

    print("--- Режим отладки парсера координат ---")
    print("Введите строку для парсинга. Для выхода введите 'exit' или 'quit'.")

    while True:
        input_string = input("> ")
        if input_string.lower() in ["exit", "quit"]:
            break

        if not input_string:
            continue

        logger.info(f"--- Начало парсинга строки: '{input_string}' ---")
        try:
            coords = parse_coordinates(input_string)
            reason = None
        except ParseError as e:
            coords = []
            reason = str(e)
        print("\n--- Итоговый результат ---")

        if reason:
            print(f"Ошибка: {reason}")
        elif not coords:
            print("Координаты не найдены или являются нулевыми.")
        else:
            print(f"Успешно найдено {len(coords)} координат:")
            for i, p in enumerate(coords):
                print(
                    f"  {i+1}. Имя: '{p.name}', Долгота: {p.lon}, Широта: {p.lat}")

            print(f"\nФормат для GeoBridge")
            for i, p in enumerate(coords):
                print(f"{p.lat}, {p.lon}")
        print("--------------------------\n")
